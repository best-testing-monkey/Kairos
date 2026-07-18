"""LibreOffice headless formula-parity tests for the Allocation sheet.

Per RFC allocation_sheet.md §8 and ticket E11-S13 (plus the Enabled-column /
live-redistribution update): recalculate the written XLSX/ODS sheets with
``soffice --headless``, export the resulting values to CSV, and spot-check
them against the pure-Python output of ``allocate()``.

Since the sheet's final Alloc % (column T) is now a *live redistribution*
among rows whose Enabled column (A) reads "true" (rather than a direct copy
of the Python-computed ``alloc``), the parity check is no longer
``calc_alloc == py_alloc / 100``. In the default state -- nothing manually
edited beyond the writer's own SELECTED/rejected pre-population of column A
-- "enabled" == "originally SELECTED", so the redistribution denominator
equals ``sum(alloc for r in selected rows) == result.gross_exposure_pct``,
giving:

    expected_alloc_fraction = (row["alloc"] / 100.0) / (result.gross_exposure_pct / 100.0)

This module also covers manually disabling a row via a direct cell edit
(openpyxl / odfpy) *before* invoking soffice, proving the redistribution is a
live spreadsheet formula, not a Python-side computation.

This module skips entirely when LibreOffice is not installed, so the rest of the
unit suite remains usable on machines without ``soffice``.
"""

import csv
import shutil
import subprocess
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from odf import opendocument
from odf.table import Table, TableCell, TableRow
from odf.text import P

from allocation import (
    AllocationConfig,
    Candidate,
    allocate,
    compute_derived,
    write_ods_sheet,
    write_xlsx_sheet,
    _sorted_for_sheet,
    _DATA_START_ROW,
    _HEADER_ROW,
)


# Skip the whole module if LibreOffice is unavailable.
SOFFICE = shutil.which("soffice") or shutil.which("libreoffice")
if not SOFFICE:
    pytest.skip(
        "LibreOffice (soffice) not found; skipping formula parity tests",
        allow_module_level=True,
    )


# CSV column indices (0-based) for the columns this module inspects.
# A=0 (Enabled), T=19 (Alloc %, the live redistribution), Q=16 (EV total).
COL_ENABLED = 0
COL_TICKER = 1
COL_STRATEGY = 3
COL_ENTRY = 5
COL_EV_TOTAL = 16
COL_ALLOC_PCT = 19


# ---------------------------------------------------------------------------
# Fixture: RFC §7 worked-example rows
# ---------------------------------------------------------------------------

def _candidate(ticker, strategy, direction, entry, stop, target, ev_pct,
               base_win_rate, n):
    """Build a Candidate with the mandatory fields populated."""
    return Candidate(
        strategy=strategy,
        ticker=ticker,
        direction=direction,
        entry=entry,
        stop=stop,
        target=target,
        ev_pct=ev_pct,
        base_win_rate=base_win_rate,
        n=n,
        backtest_period="2023-01-01 to 2023-12-31",
        sharpe=1.0,
        advised_liquidity_pct=10.0,
    )


@pytest.fixture
def rfc7_candidates():
    """Return the five worked-example candidates from RFC §7."""
    return [
        _candidate("NG=F", "close_direction", "long", 2.95, 2.97, 3.14, 1.45, 0.52, 109),
        _candidate("NG=F", "open_gap", "long", 2.95, 2.97, 3.14, 1.45, 0.52, 79),
        _candidate("REMX", "path_execution", "short", 79.73, 84.51, 73.71, 4.04, 0.47, 161),
        _candidate("V", "path_execution", "short", 200.0, 203.2, 203.6, 1.5, 0.48, 319),
        _candidate("CRM", "path_execution", "long", 200.0, 197.0, 204.6, 2.0, 0.55, 79),
    ]


@pytest.fixture
def config():
    """Default AllocationConfig with a static cluster map."""
    return AllocationConfig(
        n0=100,
        min_n=50,
        round_trip_cost_pct=0.15,
        kelly_mult=0.35,
        top_k=12,
        max_pos_pct=15.0,
        max_cluster_pct=25.0,
        gross_cap_pct=100.0,
        cluster_map={"REMX": "metals_miners", "V": "tech", "CRM": "tech"},
    )


# ---------------------------------------------------------------------------
# Manual-disable fixture: candidates solved to hit exact Kelly fractions,
# reproducing the user's literal worked example (2.85% / 1.61% -> 63.9% / 36.1%).
# ---------------------------------------------------------------------------

def _solve_base_win_rate(target_kelly_frac, kelly_mult, shrink):
    """Invert the symmetric-payoff (b=1) Kelly formula for base_win_rate.

    With entry/stop/target symmetric (risk_pct == reward_pct, so the
    geometry-fallback payoff ratio b == 1), kelly_raw simplifies to
    ``2 * p_shrunk - 1``. Solving for base_win_rate lets us hand-pick a
    Candidate whose Python- *and* sheet-computed kelly_frac (and therefore
    pos_capped_alloc / base_alloc_pct, absent any caps) equals
    ``target_kelly_frac`` (a 0-1 fraction) exactly.
    """
    kelly_raw = target_kelly_frac / kelly_mult
    p_shrunk = (kelly_raw + 1) / 2
    return 0.5 + (p_shrunk - 0.5) / shrink


def _make_symmetric_candidate(ticker, target_kelly_frac, kelly_mult, n=100, n0=100):
    """Build a Candidate with symmetric entry/stop/target (b=1) whose
    resulting kelly_frac equals ``target_kelly_frac`` exactly (verified
    against compute_derived() below, before any soffice invocation)."""
    shrink = n / (n + n0)
    base_win_rate = _solve_base_win_rate(target_kelly_frac, kelly_mult, shrink)
    return Candidate(
        strategy="manual_disable_fixture",
        ticker=ticker,
        direction="long",
        entry=100.0,
        stop=95.0,
        target=105.0,
        ev_pct=5.0,
        base_win_rate=base_win_rate,
        n=n,
        backtest_period="2023-01-01 to 2023-12-31",
        sharpe=1.0,
        advised_liquidity_pct=10.0,
    )


@pytest.fixture
def manual_disable_config():
    """Config with no position/cluster/gross caps triggered for the
    manual-disable fixture (targets are well under every cap)."""
    return AllocationConfig()  # all defaults: kelly_mult=0.35, max_pos_pct=15, ...


@pytest.fixture
def manual_disable_candidates(manual_disable_config):
    """Three SELECTED candidates: SIG_A=2.85%, SIG_B=1.61%, SIG_C=5.0% Kelly
    fraction (pre-redistribution), matching the user's literal example
    (SIG_A/SIG_B) plus a third row (SIG_C) to disable for the uneven-ratio
    sub-case."""
    km = manual_disable_config.kelly_mult
    return [
        _make_symmetric_candidate("SIG_A", 0.0285, km),
        _make_symmetric_candidate("SIG_B", 0.0161, km),
        _make_symmetric_candidate("SIG_C", 0.05, km),
    ]


def test_manual_disable_fixture_hits_exact_targets(manual_disable_candidates, manual_disable_config):
    """Sanity-check (no soffice) that the solved candidates hit the exact
    Kelly fractions used throughout this module, before any recalculation."""
    targets = {"SIG_A": 0.0285, "SIG_B": 0.0161, "SIG_C": 0.05}
    for c in manual_disable_candidates:
        derived = compute_derived(c, manual_disable_config)
        assert abs(derived["kelly_frac"] - targets[c.ticker]) < 1e-9

    result = allocate(manual_disable_candidates, manual_disable_config, {})
    selected = {r["ticker"]: r["alloc"] for r in result.rows if r["status"] == "SELECTED"}
    assert selected == {"SIG_A": pytest.approx(2.85), "SIG_B": pytest.approx(1.61), "SIG_C": pytest.approx(5.0)}


# ---------------------------------------------------------------------------
# soffice helpers
# ---------------------------------------------------------------------------

def _recalculate_to_csv(workbook_path: Path) -> list[list[str]]:
    """Recalculate ``workbook_path`` with LibreOffice and return CSV rows."""
    tmpdir = tempfile.mkdtemp()
    profile_dir = tempfile.mkdtemp()
    try:
        cmd = [
            SOFFICE,
            "--headless",
            # Use a private user profile so this invocation never collides
            # with an already-running (interactive or concurrent test)
            # soffice instance sharing the default profile, which otherwise
            # causes silent conversion failures.
            f"-env:UserInstallation=file://{profile_dir}",
            "--convert-to",
            "csv:Text - txt - csv (StarCalc):44,34,0,1",
            "--outdir",
            tmpdir,
            str(workbook_path),
        ]
        result = subprocess.run(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=False,
        )
        if result.returncode != 0:
            raise RuntimeError(
                f"soffice conversion failed: {result.stderr}\n{result.stdout}"
            )

        csv_path = Path(tmpdir) / (workbook_path.stem + ".csv")
        if not csv_path.exists():
            raise RuntimeError(f"soffice did not produce {csv_path}")

        with csv_path.open("r", newline="", encoding="utf-8") as f:
            return list(csv.reader(f))
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)
        shutil.rmtree(profile_dir, ignore_errors=True)


def _csv_float(value):
    if value is None or value == "":
        return None
    if isinstance(value, str) and value.endswith("%"):
        # Percent-formatted cells are exported as e.g. "2.4%"; the sheet's
        # underlying value is the 0-1 fraction, so divide back down.
        try:
            return float(value[:-1]) / 100.0
        except ValueError:
            return value
    try:
        return float(value)
    except ValueError:
        return value


def _write_result(result, config, fmt, tmpdir: Path) -> Path:
    """Write ``result`` to a fresh workbook/document under ``tmpdir``."""
    if fmt == "xlsx":
        wb = Workbook()
        # Drop the default "Sheet" so the workbook has exactly one sheet
        # ("Allocation"); soffice's CSV export filter emits one CSV per
        # sheet whenever a workbook has more than one sheet, which would
        # otherwise break the single-file assumption in _recalculate_to_csv.
        wb.remove(wb.active)
        write_xlsx_sheet(
            wb, result, config,
            report_date="2026-07-13", generator_version="parity-test",
        )
        path = tmpdir / "allocation.xlsx"
        wb.save(path)
    else:
        doc = opendocument.OpenDocumentSpreadsheet()
        write_ods_sheet(
            doc, result, config,
            report_date="2026-07-13", generator_version="parity-test",
        )
        path = tmpdir / "allocation.ods"
        doc.save(path)
    return path


def _set_enabled_cell_xlsx(path: Path, excel_row: int, value: str) -> None:
    """Directly edit the Enabled cell (column A) of an already-written XLSX
    file on disk, before recalculation -- simulates a user manually toggling
    the plain-text Enabled cell."""
    wb = load_workbook(path)
    ws = wb["Allocation"]
    ws.cell(row=excel_row, column=1, value=value)
    wb.save(path)


def _set_enabled_cell_ods(path: Path, data_row_index: int, value: str) -> None:
    """Directly edit the Enabled cell (column A) of an already-written ODS
    file on disk, before recalculation."""
    doc = opendocument.load(path)
    table = next(
        t for t in doc.spreadsheet.getElementsByType(Table)
        if t.getAttribute("name") == "Allocation"
    )
    rows = list(table.getElementsByType(TableRow))
    row = rows[data_row_index]
    cell = list(row.getElementsByType(TableCell))[0]
    for child in list(cell.childNodes):
        cell.removeChild(child)
    cell.addElement(P(text=value))
    cell.setAttribute("valuetype", "string")
    doc.save(path)


def _disable_tickers_and_recalculate(result, config, fmt, tickers_to_disable):
    """Write ``result``, flip the Enabled cell to "false" for each ticker in
    ``tickers_to_disable`` via a direct cell edit, then recalculate."""
    sorted_rows = _sorted_for_sheet(result.rows)
    ticker_to_offset = {row["ticker"]: i for i, row in enumerate(sorted_rows)}

    tmpdir = tempfile.mkdtemp()
    try:
        tmpdir = Path(tmpdir)
        path = _write_result(result, config, fmt, tmpdir)

        for ticker in tickers_to_disable:
            offset = ticker_to_offset[ticker]
            if fmt == "xlsx":
                excel_row = _DATA_START_ROW + 1 + offset
                _set_enabled_cell_xlsx(path, excel_row, "false")
            else:
                data_row_index = _DATA_START_ROW + offset
                _set_enabled_cell_ods(path, data_row_index, "false")

        csv_rows = _recalculate_to_csv(path)
        return sorted_rows, csv_rows
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestLibreOfficeParity:
    """Formula parity between allocate() and LibreOffice-recalculated sheets.

    Covers scenarios 1 & 2 from the plan (default Enabled-column population,
    and default live redistribution) across many enabled_mask combinations,
    since in the default state (nothing manually toggled) "enabled" ==
    "originally SELECTED": the writer pre-populates column A as "true" for
    SELECTED rows and "false" for everything else, so these Python-level
    enabled_mask variations exercise many different selected-row-count
    scenarios "for free".
    """

    def _write_and_recalculate(self, candidates, config, enabled_mask, fmt):
        """Allocate, write to temp file, recalculate, return (result, csv_rows)."""
        result = allocate(candidates, config, enabled_mask or {})
        tmpdir = tempfile.mkdtemp()
        try:
            tmpdir = Path(tmpdir)
            path = _write_result(result, config, fmt, tmpdir)
            csv_rows = _recalculate_to_csv(path)
            return result, csv_rows
        finally:
            shutil.rmtree(tmpdir, ignore_errors=True)

    def _check_data_rows_and_redistribution(self, result, csv_rows, fmt):
        """Spot-check the Enabled column population and the live
        redistribution formula (column T) against the plan's expected-value
        formula for the default state."""
        # Data rows begin at CSV index _HEADER_ROW (spreadsheet row
        # _HEADER_ROW + 1), in the same EV-total-descending order the
        # writers use for the sheet.
        sorted_rows = _sorted_for_sheet(result.rows)
        data_csv = csv_rows[_HEADER_ROW:]
        assert len(data_csv) >= len(sorted_rows), (
            f"{fmt}: expected at least {len(sorted_rows)} data rows, "
            f"got {len(data_csv)}"
        )

        gross = result.gross_exposure_pct
        t_sum = 0.0

        for offset, row in enumerate(sorted_rows):
            excel_row = _DATA_START_ROW + 1 + offset
            csv_row = data_csv[offset]

            # Scenario 1: Enabled column A defaults to "true" for SELECTED
            # rows, "false" for everything else (rejected candidates).
            expected_enabled = "true" if row.get("status") == "SELECTED" else "false"
            assert csv_row[COL_ENABLED] == expected_enabled, (
                f"{fmt} A{excel_row}: expected Enabled={expected_enabled!r}, "
                f"got {csv_row[COL_ENABLED]!r}"
            )

            assert csv_row[COL_TICKER] == row.get("ticker", ""), (
                f"{fmt} B{excel_row} ticker mismatch"
            )
            assert csv_row[COL_STRATEGY] == row.get("strategy", ""), (
                f"{fmt} D{excel_row} strategy mismatch"
            )
            expected_entry = row.get("entry")
            if expected_entry is not None:
                assert abs(_csv_float(csv_row[COL_ENTRY]) - expected_entry) <= 1e-9, (
                    f"{fmt} F{excel_row} entry mismatch"
                )

            calc_t = _csv_float(csv_row[COL_ALLOC_PCT]) or 0.0
            t_sum += calc_t

            # Scenario 2: default live redistribution. Since "enabled" ==
            # "originally SELECTED" in the default state, the redistribution
            # denominator equals sum(alloc for SELECTED rows) ==
            # result.gross_exposure_pct.
            if row.get("status") == "SELECTED":
                py_alloc = row.get("alloc")
                assert py_alloc is not None and gross, (
                    f"{fmt} T{excel_row}: missing alloc/gross_exposure_pct for parity check"
                )
                expected_alloc_fraction = (py_alloc / 100.0) / (gross / 100.0)
                assert abs(calc_t - expected_alloc_fraction) <= 5e-4, (
                    f"{fmt} T{excel_row}: expected redistributed alloc "
                    f"{expected_alloc_fraction}, got {calc_t}"
                )

                # Column Q is EV total = ev_pct/100 * Alloc % (redistributed).
                ev_pct = row.get("ev_pct")
                if ev_pct is not None:
                    expected_ev_total = (ev_pct / 100.0) * expected_alloc_fraction
                    calc_ev_total = _csv_float(csv_row[COL_EV_TOTAL])
                    assert calc_ev_total is not None, f"{fmt} Q{excel_row}: missing EV total value"
                    assert abs(expected_ev_total - calc_ev_total) <= 5e-4, (
                        f"{fmt} Q{excel_row}: expected EV total {expected_ev_total}, "
                        f"got {calc_ev_total}"
                    )
            else:
                # Disabled/rejected rows must contribute exactly 0 to Alloc %.
                assert abs(calc_t) <= 5e-4, (
                    f"{fmt} T{excel_row}: expected 0 for non-SELECTED row, got {calc_t}"
                )

        # The redistributed Alloc % column must sum to 100% whenever at
        # least one row is selected (renormalization guarantee).
        if result.selected_count > 0:
            assert abs(t_sum - 1.0) <= 5e-3, (
                f"{fmt}: expected SUM(Alloc %) == 1.0 across all rows, got {t_sum}"
            )

    @pytest.mark.parametrize("fmt", ["xlsx", "ods"])
    def test_all_enabled_mask(self, rfc7_candidates, config, fmt):
        """All-enabled mask: conversion succeeds and default redistribution matches."""
        result, csv_rows = self._write_and_recalculate(
            rfc7_candidates, config, {}, fmt
        )
        self._check_data_rows_and_redistribution(result, csv_rows, fmt)

    @pytest.mark.parametrize("ticker", ["NG=F", "REMX", "V", "CRM"])
    @pytest.mark.parametrize("fmt", ["xlsx", "ods"])
    def test_single_signal_disabled(self, rfc7_candidates, config, ticker, fmt):
        """Disable exactly one ticker at a time (Python-level enabled_mask,
        i.e. before selection even happens) and verify parity."""
        enabled_mask = {c.ticker: True for c in rfc7_candidates}
        enabled_mask[ticker] = False
        result, csv_rows = self._write_and_recalculate(
            rfc7_candidates, config, enabled_mask, fmt
        )
        self._check_data_rows_and_redistribution(result, csv_rows, fmt)

    @pytest.mark.parametrize("fmt", ["xlsx", "ods"])
    def test_random_masks(self, rfc7_candidates, config, fmt):
        """100 seeded random enabled/disabled masks."""
        import random
        rng = random.Random(42)
        tickers = list({c.ticker for c in rfc7_candidates})

        for _ in range(100):
            enabled_mask = {t: rng.choice([True, False]) for t in tickers}
            result, csv_rows = self._write_and_recalculate(
                rfc7_candidates, config, enabled_mask, fmt
            )
            self._check_data_rows_and_redistribution(result, csv_rows, fmt)


class TestBlankRowAndInstruction:
    """Scenario 4: row 18 is genuinely blank and the instruction text lives
    at row 19 after recalculation."""

    @pytest.mark.parametrize("fmt", ["xlsx", "ods"])
    def test_blank_row_18_and_instruction_row_19(self, rfc7_candidates, config, fmt):
        result = allocate(rfc7_candidates, config, {})
        tmpdir = tempfile.mkdtemp()
        try:
            path = _write_result(result, config, fmt, Path(tmpdir))
            csv_rows = _recalculate_to_csv(path)
        finally:
            shutil.rmtree(tmpdir, ignore_errors=True)

        # CSV row index 17 == spreadsheet row 18 (0-based).
        blank_row = csv_rows[17]
        assert all(cell == "" for cell in blank_row), (
            f"{fmt}: expected row 18 to be blank, got {blank_row}"
        )

        instruction_row = csv_rows[18]
        assert instruction_row[0] != "", f"{fmt}: expected instruction text at row 19"
        assert "column A" in instruction_row[0]
        assert "computed" in instruction_row[0]


class TestManualDisableRedistribution:
    """Scenario 3: manually disable a row via a direct cell edit (openpyxl /
    odfpy) *before* invoking soffice, proving the redistribution is a live
    spreadsheet formula. Reproduces the user's literal worked example:
    two enabled signals with original allocs 2.85%/1.61% -> 63.9%/36.1%,
    and a single remaining enabled signal -> 100%.
    """

    @pytest.mark.parametrize("fmt", ["xlsx", "ods"])
    def test_disable_one_of_three_uneven_ratio(
        self, manual_disable_candidates, manual_disable_config, fmt
    ):
        """Disable SIG_C (5.0% base alloc), leaving SIG_A (2.85%) and
        SIG_B (1.61%) enabled -> 63.9% / 36.1% (the user's literal example)."""
        result = allocate(manual_disable_candidates, manual_disable_config, {})
        assert result.selected_count == 3

        sorted_rows, csv_rows = _disable_tickers_and_recalculate(
            result, manual_disable_config, fmt, tickers_to_disable=["SIG_C"]
        )
        data_csv = csv_rows[_HEADER_ROW:]

        alloc_by_ticker = {}
        for offset, row in enumerate(sorted_rows):
            calc_t = _csv_float(data_csv[offset][COL_ALLOC_PCT])
            alloc_by_ticker[row["ticker"]] = calc_t

        assert alloc_by_ticker["SIG_C"] == pytest.approx(0.0, abs=5e-4)
        assert alloc_by_ticker["SIG_A"] == pytest.approx(0.639, abs=5e-4)
        assert alloc_by_ticker["SIG_B"] == pytest.approx(0.361, abs=5e-4)
        assert (alloc_by_ticker["SIG_A"] + alloc_by_ticker["SIG_B"]) == pytest.approx(1.0, abs=5e-3)

    @pytest.mark.parametrize("fmt", ["xlsx", "ods"])
    def test_disable_two_of_three_single_signal_full_alloc(
        self, manual_disable_candidates, manual_disable_config, fmt
    ):
        """Disable SIG_B and SIG_C, leaving only SIG_A enabled -> 100%."""
        result = allocate(manual_disable_candidates, manual_disable_config, {})
        assert result.selected_count == 3

        sorted_rows, csv_rows = _disable_tickers_and_recalculate(
            result, manual_disable_config, fmt, tickers_to_disable=["SIG_B", "SIG_C"]
        )
        data_csv = csv_rows[_HEADER_ROW:]

        alloc_by_ticker = {}
        for offset, row in enumerate(sorted_rows):
            calc_t = _csv_float(data_csv[offset][COL_ALLOC_PCT])
            alloc_by_ticker[row["ticker"]] = calc_t

        assert alloc_by_ticker["SIG_B"] == pytest.approx(0.0, abs=5e-4)
        assert alloc_by_ticker["SIG_C"] == pytest.approx(0.0, abs=5e-4)
        assert alloc_by_ticker["SIG_A"] == pytest.approx(1.0, abs=5e-3)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
