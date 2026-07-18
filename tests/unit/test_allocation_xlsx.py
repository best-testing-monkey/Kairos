"""Test suite for E11-S09: write_xlsx_sheet() XLSX allocation sheet writer.

Uses an in-memory openpyxl workbook, saves to BytesIO, reloads, and asserts the
sheet layout, formulas, config block, and protection settings per the RFC and
ticket acceptance criteria.
"""

from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from allocation import (
    AllocationConfig,
    AllocationResult,
    Candidate,
    compute_derived,
    render_formula,
    write_xlsx_sheet,
    _sorted_for_sheet,
)


def _make_candidate(**kwargs):
    """Build a Candidate with sensible defaults; override via kwargs."""
    defaults = {
        "strategy": "path_execution",
        "ticker": "TEST",
        "direction": "long",
        "entry": 100.0,
        "stop": 95.0,
        "target": 110.0,
        "ev_pct": 5.0,
        "base_win_rate": 0.55,
        "n": 100,
        "backtest_period": "2023-01-01 to 2023-12-31",
        "sharpe": 1.2,
        "advised_liquidity_pct": 10.0,
    }
    defaults.update(kwargs)
    return Candidate(**defaults)


def _make_row(candidate, config, status, flags=None, alloc=None):
    """Construct a result row dict from a Candidate."""
    derived = compute_derived(candidate, config)
    row = {
        "strategy": candidate.strategy,
        "ticker": candidate.ticker,
        "direction": candidate.direction,
        "entry": candidate.entry,
        "stop": candidate.stop,
        "target": candidate.target,
        "ev_pct": candidate.ev_pct,
        "base_win_rate": candidate.base_win_rate,
        "n": candidate.n,
        "backtest_period": candidate.backtest_period,
        "sharpe": candidate.sharpe,
        "advised_liquidity_pct": candidate.advised_liquidity_pct,
        "derived": derived,
        "status": status,
        "flags": list(flags or []),
    }
    if alloc is not None:
        row["alloc"] = alloc
    return row


@pytest.fixture
def config():
    return AllocationConfig(
        cluster_map={"REMX": "metals_miners", "NG": "energy"},
    )


@pytest.fixture
def result(config):
    """AllocationResult with one selected and one rejected candidate."""
    selected = _make_candidate(
        ticker="REMX", direction="short", entry=79.73, stop=84.51, target=73.71,
        ev_pct=4.04, base_win_rate=0.47, n=161,
    )
    selected_derived = compute_derived(selected, config)
    selected_alloc = min(selected_derived["kelly_frac"] * 100, config.max_pos_pct)

    rejected = _make_candidate(
        ticker="NG", direction="long", entry=3.0, stop=2.97, target=3.14,
        ev_pct=1.45, base_win_rate=0.52, n=109,
    )

    rows = [
        _make_row(selected, config, "SELECTED", ["DATA_MISMATCH"], selected_alloc),
        _make_row(rejected, config, "BELOW_TOPK", []),
    ]
    return AllocationResult(
        rows=rows,
        selected_count=1,
        gross_exposure_pct=selected_alloc,
        rejection_counts={"BELOW_TOPK": 1},
    )


@pytest.fixture
def workbook(result, config):
    """Return a loaded openpyxl workbook containing the Allocation sheet."""
    wb = Workbook()
    write_xlsx_sheet(
        wb,
        result,
        config,
        report_date="2026-07-13",
        generator_version="1.0.0-test",
    )
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return load_workbook(buffer)


class TestSheetStructure:
    """High-level layout checks."""

    def test_sheet_name(self, workbook):
        assert "Allocation" in workbook.sheetnames
        ws = workbook["Allocation"]
        assert ws.title == "Allocation"

    def test_title_row(self, workbook):
        ws = workbook["Allocation"]
        assert "Portfolio Allocation" in str(ws["A1"].value)
        assert "2026-07-13" in str(ws["B1"].value)
        assert "1.0.0-test" in str(ws["C1"].value)

    def test_config_block_layout(self, workbook, config):
        ws = workbook["Allocation"]
        # The editable value cells are E3:E13, one row per AllocationConfig field.
        assert ws["D3"].value == "n0"
        assert ws["E3"].value == config.n0
        assert ws["F3"].value == 100  # shipped default

        assert ws["D4"].value == "round_trip_cost_pct"
        assert abs(ws["E4"].value - config.round_trip_cost_pct / 100.0) < 1e-9
        assert abs(ws["F4"].value - 0.0015) < 1e-9

        assert ws["D9"].value == "equity"
        assert ws["E9"].value == "" or ws["E9"].value is None

        # All 11 config rows are present (rows 3 through 13).
        for row in range(3, 14):
            assert ws.cell(row=row, column=4).value is not None

    def test_summary_block_formulas(self, workbook):
        ws = workbook["Allocation"]
        assert ws["A14"].value == "Selected count"
        assert str(ws["D14"].value).startswith("=")
        assert ws["A15"].value == "Gross exposure %"
        assert str(ws["D15"].value).startswith("=")
        assert ws["A16"].value == "Enabled count"
        assert str(ws["D16"].value).startswith("=")
        assert ws["E14"].value == render_formula("gross_scale", 14, "xlsx")

    def test_blank_row_18(self, workbook):
        ws = workbook["Allocation"]
        for col_idx in range(1, 20):
            assert ws.cell(row=18, column=col_idx).value in (None, "")

    def test_instruction_line(self, workbook):
        ws = workbook["Allocation"]
        assert ws["A19"].value is not None
        assert isinstance(ws["A19"].value, str)
        assert "column A" in ws["A19"].value


class TestHeaderRow:
    """Header row 20 spot checks per RFC §5.2."""

    def test_header_row_static_columns(self, workbook):
        ws = workbook["Allocation"]
        assert ws["A20"].value == "Enabled"
        assert ws["B20"].value == "Ticker"
        assert ws["C20"].value == "Cluster"
        assert ws["D20"].value == "Strategy"
        assert ws["E20"].value == "Dir"
        assert ws["F20"].value == "Entry"
        assert ws["G20"].value == "Stop"
        assert ws["H20"].value == "Target"
        assert ws["O20"].value == "EV raw %"

    def test_header_row_formula_columns(self, workbook):
        ws = workbook["Allocation"]
        assert ws["P20"].value == "EV net %"
        assert ws["Q20"].value == "EV total"
        assert ws["T20"].value == "Alloc %"
        assert ws["V20"].value == "Flags"
        assert ws["W20"].value == "Advised liq % (ignored)"

    def test_header_row_helper_columns(self, workbook):
        ws = workbook["Allocation"]
        assert ws["Z20"].value == "b"
        assert ws["AB20"].value == "shrink"
        assert ws["AL20"].value == "cluster_scale"
        assert ws["AM20"].value == "enabled_flag"
        assert ws["AN20"].value == "base_alloc_pct"


class TestDataRows:
    """Checks for the per-candidate data rows."""

    def test_data_rows_written_in_sorted_order(self, workbook, result):
        ws = workbook["Allocation"]
        sorted_rows = _sorted_for_sheet(result.rows)
        assert ws["B21"].value == sorted_rows[0]["ticker"]
        assert ws["B22"].value == sorted_rows[1]["ticker"]

    def test_static_values_from_result(self, workbook, result):
        ws = workbook["Allocation"]
        row0 = _sorted_for_sheet(result.rows)[0]
        assert ws["A21"].value == "true"  # SELECTED row -> Enabled defaults to "true"
        assert ws["B21"].value == row0["ticker"]
        assert ws["C21"].value == "metals_miners"
        assert ws["D21"].value == row0["strategy"]
        assert ws["E21"].value == "Short"
        assert ws["F21"].value == row0["entry"]
        assert ws["G21"].value == row0["stop"]
        assert ws["H21"].value == row0["target"]
        assert abs(ws["O21"].value - row0["ev_pct"] / 100.0) < 1e-9

    def test_rejected_row_enabled_defaults_false(self, workbook, result):
        ws = workbook["Allocation"]
        row1 = _sorted_for_sheet(result.rows)[1]
        assert row1["status"] == "BELOW_TOPK"
        assert ws["A22"].value == "false"

    def test_formula_cell_matches_render_formula(self, workbook):
        ws = workbook["Allocation"]
        assert ws["P21"].value == render_formula("P", 21, "xlsx")
        assert ws["T21"].value == render_formula("T", 21, "xlsx")
        assert ws["AL21"].value == render_formula("AL", 21, "xlsx")
        assert ws["AM21"].value == render_formula("AM", 21, "xlsx")
        assert ws["AN21"].value == render_formula("AN", 21, "xlsx")


class TestProtection:
    """Sheet protection checks."""

    def test_protection_enabled(self, workbook):
        ws = workbook["Allocation"]
        assert ws.protection.sheet is True

    def test_editable_config_cells_unlocked(self, workbook):
        ws = workbook["Allocation"]
        for row in range(3, 14):
            assert ws.cell(row=row, column=5).protection.locked is False

    def test_column_a_unlocked(self, workbook):
        ws = workbook["Allocation"]
        for row in range(1, ws.max_row + 1):
            assert ws.cell(row=row, column=1).protection.locked is False

    def test_column_o_unlocked(self, workbook):
        ws = workbook["Allocation"]
        for row in range(1, ws.max_row + 1):
            assert ws.cell(row=row, column=15).protection.locked is False

    def test_other_cells_locked(self, workbook):
        ws = workbook["Allocation"]
        assert ws["B1"].protection.locked is True
        assert ws["D3"].protection.locked is True
        assert ws["F3"].protection.locked is True
        assert ws["P21"].protection.locked is True


class TestClusterExposure:
    """Section B cluster-exposure table checks."""

    def test_cluster_header_and_rows(self, workbook, config):
        ws = workbook["Allocation"]
        # The cluster block starts two rows below the data.
        assert ws["A24"].value == "Cluster"
        assert ws["B24"].value == "Positions"
        assert ws["C24"].value == "Gross %"
        assert ws["D24"].value == "Cap %"
        assert ws["E24"].value == "Capped?"

        clusters = {ws.cell(row=r, column=1).value for r in range(25, 27)}
        assert clusters == {"energy", "metals_miners"}


class TestGrossCapScaling:
    """Verify the sheet formulas scale gross exposure down to the cap."""

    def test_gross_scale_formula_uses_post_cluster_total(self):
        """If raw position-capped allocs exceed gross_cap_pct, gross_scale < 1."""
        candidates = [
            _make_candidate(
                ticker=f"T{i:02d}",
                entry=100.0,
                stop=95.0,
                target=110.0,
                ev_pct=5.0,
                base_win_rate=0.60,
                n=200,
            )
            for i in range(12)
        ]
        config = AllocationConfig(
            top_k=12,
            max_pos_pct=15.0,
            gross_cap_pct=100.0,
        )
        result = AllocationResult(
            rows=[
                _make_row(
                    c,
                    config,
                    "SELECTED",
                    alloc=min(compute_derived(c, config)["kelly_frac"] * 100, 15.0),
                )
                for c in candidates
            ],
            selected_count=12,
            gross_exposure_pct=100.0,
            rejection_counts={},
        )

        wb = Workbook()
        write_xlsx_sheet(
            wb,
            result,
            config,
            report_date="2026-07-13",
            generator_version="cap-test",
        )
        ws = wb["Allocation"]

        # The gross_scale formula must reference AG and AL via SUMPRODUCT, not
        # just SUM(AL). If it only summed AL, the scale factor would be 1.
        gross_scale_formula = str(ws["E14"].value)
        assert "SUMPRODUCT" in gross_scale_formula
        assert "AG$21:AG$401" in gross_scale_formula
        assert "AL$21:AL$401" in gross_scale_formula

        # Simulate what the formula evaluates to with no cluster caps.
        # AG = position-capped alloc %, AL = 1 for each row.
        ag_values = [min(compute_derived(c, config)["kelly_frac"] * 100, 15.0)
                     for c in candidates]
        post_cluster_total = sum(ag_values)  # no cluster caps
        assert post_cluster_total > config.gross_cap_pct
        expected_scale = config.gross_cap_pct / post_cluster_total
        assert expected_scale < 1.0

        # Each AN cell (base_alloc_pct, the old alloc_pct formula relocated)
        # should contain AG*AL*gross_scale.
        for offset in range(len(candidates)):
            excel_row = 21 + offset
            an_formula = str(ws[f"AN{excel_row}"].value)
            assert f"AG{excel_row}" in an_formula
            assert f"AL{excel_row}" in an_formula
            assert "$E$14" in an_formula


class TestPercentFormatting:
    """Percent-labeled cells use 0-1 fraction values with 0.0% number format."""

    def test_percent_number_format_on_data_columns(self, workbook):
        ws = workbook["Allocation"]
        for col in ("I", "J", "O", "P", "Q", "T", "X", "Y", "AA", "AF", "AN"):
            assert ws[f"{col}21"].number_format == "0.0%"

    def test_summary_cells_percent_format(self, workbook):
        ws = workbook["Allocation"]
        assert ws["D15"].number_format == "0.0%"
        assert ws["D17"].number_format == "0.0%"


class TestAutofilter:
    """Autofilter is enabled over the header + data range."""

    def test_auto_filter_ref(self, workbook, result):
        ws = workbook["Allocation"]
        data_end_row = 20 + len(result.rows)
        assert ws.auto_filter.ref == f"A20:AN{data_end_row}"


class TestEvTotalSummaryRow:
    """Row 17 carries the EV total summary formula, percent-formatted."""

    def test_ev_total_summary_row(self, workbook):
        ws = workbook["Allocation"]
        assert ws["A17"].value == "EV total"
        assert str(ws["D17"].value).startswith("=")
        assert "Q21:Q401" in str(ws["D17"].value)


class TestEvTotalColumn:
    """'EV total' column Q: ev_pct fraction * alloc fraction (redistributed)."""

    def test_ev_total_formula(self, workbook):
        ws = workbook["Allocation"]
        assert ws["Q21"].value == render_formula("Q", 21, "xlsx")
        assert ws["Q21"].value == "=O21*T21"


class TestAllocPctRedistribution:
    """New column T: live redistribution formula among enabled rows."""

    def test_alloc_pct_formula_structure(self, workbook):
        ws = workbook["Allocation"]
        formula = str(ws["T21"].value)
        assert "AM21" in formula
        assert "AN21" in formula
        assert "AN$21:AN$401" in formula
        assert "AM$21:AM$401" in formula
        assert "SUMPRODUCT" in formula


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
