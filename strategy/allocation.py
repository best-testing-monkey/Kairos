"""allocation.py — Candidate schema and fetch_signals adapter for portfolio allocation.

Adapts the existing signals-report row data (stats_rows, advice_rows from kairos_signals.py)
into structured Candidate objects matching RFC allocation_sheet.md §3.

Implements AllocationConfig (RFC §3.1 defaults) and compute_derived (RFC §4.2 per-row formulas).
"""

import csv
import math
from dataclasses import dataclass, field
from typing import Optional

from openpyxl.styles import Protection

from kairos_signals import _ev_pct_value, format_table


@dataclass
class Candidate:
    """Structured representation of a single trade candidate for allocation.

    Maps one-to-one with a non-FLAT signal from kairos_signals.py run().
    Fields match RFC allocation_sheet.md §3 required schema, plus nullable fields.
    """
    strategy: str
    ticker: str
    direction: str  # "long" or "short" (FLAT excluded)
    entry: float
    stop: float
    target: float
    ev_pct: float
    base_win_rate: float
    n: int
    backtest_period: str
    sharpe: float
    advised_liquidity_pct: float
    avg_win_pct: Optional[float] = None
    avg_loss_pct: Optional[float] = None
    avg_holding_days: Optional[float] = None
    model: Optional[str] = None


@dataclass
class AllocationConfig:
    """Configuration for portfolio allocation sizing and gating.

    Per RFC allocation_sheet.md §3.1, defaults are deliberately round numbers
    swept in Phantom Ledger. Do not ship precise-looking fitted values.
    """
    n0: int = 100  # Shrinkage constant; weight of the "no edge" prior
    min_n: int = 50  # Reject signals with fewer backtest trades
    round_trip_cost_pct: float = 0.15  # Assumed total cost per round trip
    kelly_mult: float = 0.35  # Fractional Kelly multiplier
    top_k: int = 12  # Max number of positions
    max_pos_pct: float = 15  # Cap per position, % of equity
    max_cluster_pct: float = 25  # Cap per correlation cluster, % of equity
    gross_cap_pct: float = 100  # Total gross exposure cap
    dust_min_pct: float = 1.0  # Zero out final allocations below this, % of equity
    equity: Optional[float] = None  # Optional account equity for currency-amount column
    cluster_map: dict = field(default_factory=dict)  # ticker -> cluster name, static mapping


@dataclass
class AllocationResult:
    """Result of allocate() orchestration per RFC allocation_sheet.md §4.4.

    The top-level output dataclass carrying the full allocation decision, including
    all rows (selected and rejected), summary statistics, and rejection counts.
    """
    rows: list[dict] = field(default_factory=list)  # One per input candidate with all fields + derived/status/flags/alloc
    selected_count: int = 0  # Number of rows with status == "SELECTED"
    gross_exposure_pct: float = 0.0  # Sum of alloc across status == "SELECTED" rows
    rejection_counts: dict[str, int] = field(default_factory=dict)  # status -> count for all non-selected rows


def compute_ev_ratio(c: Candidate, derived: dict) -> tuple[float, bool]:
    """Compute ev_ratio and DATA_MISMATCH flag per RFC allocation_sheet.md §4.3.

    Implements the data-quality check comparing empirical EV (from backtest) against
    EV implied by geometry (base_win_rate, risk_pct, reward_pct). Flags significant
    divergence.

    Args:
        c: Candidate object with ev_pct and base_win_rate
        derived: dict from compute_derived() with keys risk_pct, reward_pct

    Returns:
        (ev_ratio, is_mismatch) tuple where:
        - ev_ratio: float, the ratio ev_pct / ev_implied (or 0.0 if ev_implied near zero)
        - is_mismatch: bool, True iff ev_ratio is outside [0.5, 2.0] and ev_implied is not near zero.
                       If ev_implied is near zero (< 1e-9 in absolute value), treat as not-mismatched
                       since the ratio is undefined, not a data problem.
    """
    risk_pct = derived["risk_pct"]
    reward_pct = derived["reward_pct"]

    # Compute ev_implied per RFC §4.3
    ev_implied = c.base_win_rate * reward_pct - (1 - c.base_win_rate) * risk_pct

    # Guard against near-zero denominator
    if abs(ev_implied) < 1e-9:
        # Not mismatched when ev_implied is too small to define a meaningful ratio
        return 0.0, False

    # Compute ev_ratio
    ev_ratio = c.ev_pct / ev_implied

    # Check if ratio is outside the band [0.5, 2.0]
    is_mismatch = ev_ratio < 0.5 or ev_ratio > 2.0

    return ev_ratio, is_mismatch


def compute_derived(c: Candidate, config: AllocationConfig) -> dict:
    """Compute per-row derived columns per RFC allocation_sheet.md §4.2.

    Implements the exact formulas for derived allocation metrics, including the two branches:
    - Empirical branch: when avg_win_pct and avg_loss_pct are both present
    - Geometry-fallback branch: when either is None, use TP/SL geometry

    Args:
        c: Candidate object with all required fields
        config: AllocationConfig with n0, round_trip_cost_pct, kelly_mult

    Returns:
        dict with keys: risk_pct, reward_pct, b, loss_pct, shrink, ev_shrunk,
        ev_net, p_shrunk, kelly_raw, kelly_frac, score

    Note: No division-by-zero checks needed. risk_pct and loss_pct are guaranteed
    positive (stop/target placement validated in schema), and shrink is in [0,1).
    """
    # Per-row derived columns, per RFC §4.2
    risk_pct = abs(c.stop - c.entry) / c.entry * 100
    reward_pct = abs(c.target - c.entry) / c.entry * 100

    # Payoff ratio: empirical when available, geometry as fallback
    if c.avg_win_pct is not None and c.avg_loss_pct is not None:
        b = c.avg_win_pct / c.avg_loss_pct
        loss_pct = c.avg_loss_pct
    else:
        b = reward_pct / risk_pct
        loss_pct = risk_pct

    shrink = c.n / (c.n + config.n0)
    ev_shrunk = c.ev_pct * shrink
    ev_net = ev_shrunk - config.round_trip_cost_pct

    p_shrunk = 0.5 + (c.base_win_rate - 0.5) * shrink
    kelly_raw = p_shrunk - (1 - p_shrunk) / b
    kelly_frac = max(kelly_raw, 0) * config.kelly_mult

    score = ev_net / loss_pct

    return {
        "risk_pct": risk_pct,
        "reward_pct": reward_pct,
        "b": b,
        "loss_pct": loss_pct,
        "shrink": shrink,
        "ev_shrunk": ev_shrunk,
        "ev_net": ev_net,
        "p_shrunk": p_shrunk,
        "kelly_raw": kelly_raw,
        "kelly_frac": kelly_frac,
        "score": score,
    }


def validate_candidate(c: Candidate) -> Optional[str]:
    """Validate a Candidate against schema constraints per RFC allocation_sheet.md §3.

    Returns "SCHEMA_ERROR" if:
      - Any required field is None
      - Any numeric required field (entry, stop, target, ev_pct, base_win_rate, sharpe)
        is non-finite (NaN or inf)
      - Direction/stop/target placement is inconsistent with direction:
        * direction == "long" requires stop < entry < target
        * direction == "short" requires target < entry < stop

    Returns None if the candidate is valid.

    Args:
        c: Candidate object to validate

    Returns:
        "SCHEMA_ERROR" if invalid, None if valid
    """
    # Check all required fields are not None
    if (c.strategy is None or c.ticker is None or c.direction is None or
        c.entry is None or c.stop is None or c.target is None or
        c.ev_pct is None or c.base_win_rate is None or c.n is None or
        c.backtest_period is None or c.sharpe is None):
        return "SCHEMA_ERROR"

    # Check that numeric fields are finite (not NaN or inf)
    numeric_fields = [c.entry, c.stop, c.target, c.ev_pct, c.base_win_rate, c.sharpe]
    for field in numeric_fields:
        if not math.isfinite(field):
            return "SCHEMA_ERROR"

    # Check direction/stop/target placement consistency
    if c.direction == "long":
        # For long: stop < entry < target
        if not (c.stop < c.entry < c.target):
            return "SCHEMA_ERROR"
    elif c.direction == "short":
        # For short: target < entry < stop
        if not (c.target < c.entry < c.stop):
            return "SCHEMA_ERROR"
    else:
        # Invalid direction value (not "long" or "short")
        return "SCHEMA_ERROR"

    return None


def fetch_signals(stats_rows, advice_rows):
    """Adapt stats_rows and advice_rows from kairos_signals.py run() into Candidate objects.

    Args:
        stats_rows: list of dicts with keys from kairos_signals.py STATS_COLUMNS
                    (strategy, symbol, direction, entry, stop, target, expected_value,
                     base_sharpe, base_win_rate, etc.)
        advice_rows: list of dicts with keys (expected_value, entry, base_win_rate,
                     base_signals, oracle_signals, signal)

    Returns:
        list of Candidate objects, one per non-FLAT stats_row, correctly paired
        to its corresponding advice_row by list index (both lists are built in lockstep
        in kairos_signals.py run()).

    Exclusions:
        - Rows with direction == "FLAT" are excluded entirely
        - Stats rows and advice rows are matched by index, so both lists must be
          the same length and built in the same order
    """
    candidates = []

    for stats_row, advice_row in zip(stats_rows, advice_rows):
        # Exclude FLAT direction rows
        direction_str = stats_row.get("direction", "").upper()
        if direction_str == "FLAT":
            continue

        # Normalize direction to lowercase "long" or "short"
        direction = direction_str.lower()

        # Extract ev_pct using the same helper as kairos_signals.py
        ev_pct = _ev_pct_value(
            stats_row.get("expected_value"),
            stats_row.get("entry")
        )
        # If ev_pct could not be computed, skip this row
        # (though in practice this should not happen for non-FLAT signals)
        if ev_pct is None:
            continue

        # Fallback for n: use base_signals, then oracle_signals, then None
        # (mirroring build_signals_table's fallback in kairos_signals.py:290-293)
        n_value = None
        base_signals = advice_row.get("base_signals")
        oracle_signals = advice_row.get("oracle_signals")

        if base_signals is not None and not _is_missing(base_signals):
            n_value = int(base_signals)
        elif oracle_signals is not None and not _is_missing(oracle_signals):
            n_value = int(oracle_signals)

        # Skip if n could not be determined
        if n_value is None:
            continue

        # Compute advised_liquidity_pct from size
        size = stats_row.get("size", 0.0)
        advised_liquidity_pct = size * 100.0 if size else 0.0

        # Construct the Candidate
        candidate = Candidate(
            strategy=stats_row.get("strategy", ""),
            ticker=stats_row.get("symbol", ""),
            direction=direction,
            entry=float(stats_row.get("entry", 0.0)),
            stop=float(stats_row.get("stop", 0.0)),
            target=float(stats_row.get("target", 0.0)),
            ev_pct=ev_pct,
            base_win_rate=float(stats_row.get("base_win_rate", 0.0)),
            n=n_value,
            backtest_period=str(stats_row.get("backtest_period", "")),
            sharpe=float(stats_row.get("base_sharpe", 0.0)),
            advised_liquidity_pct=advised_liquidity_pct,
            avg_win_pct=None,  # Not present in current data; v1 nullable
            avg_loss_pct=None,  # Not present in current data; v1 nullable
            avg_holding_days=None,  # Not present in current data; v1 nullable
            model=stats_row.get("model", ""),
        )

        candidates.append(candidate)

    return candidates


def select_candidates(candidates: list[Candidate], config: AllocationConfig, enabled_mask: dict) -> list[dict]:
    """Select and reject candidates through gating, collapse, and top-K ranking.

    Implements RFC §4.4 selection algorithm: gate → per-asset collapse → rank+top-K.

    Rejection reasons (per RFC §4.5):
    - SCHEMA_ERROR: required field missing or direction/stop/target inconsistent
    - DISABLED: enabled_mask[ticker] is False
    - LOW_N: n < config.min_n
    - NEG_EV_NET: ev_net <= 0
    - DIRECTION_CONFLICT: both long and short survive gating for same ticker
    - DUP_ASSET: duplicate ticker/direction, not the max-score row
    - BELOW_TOPK: survivor after collapse, but outside top config.top_k

    Args:
        candidates: list of Candidate objects
        config: AllocationConfig with gating/ranking parameters
        enabled_mask: dict mapping ticker -> bool; get(ticker, True) defaults to enabled

    Returns:
        list of dicts (one per candidate), in input order, each containing:
        - All original Candidate fields (strategy, ticker, direction, entry, stop, target, etc.)
        - derived: dict with keys from compute_derived (risk_pct, reward_pct, score, etc.)
        - status: rejection reason (str) or None for rows proceeding to E11-S06 sizing
        - flags: list starting with ["DATA_MISMATCH"] if ev_ratio flagged, else []
    """
    from collections import defaultdict

    # Convert all candidates to dicts and prepare for processing
    output_rows = {}  # original_index -> row dict

    for i, c in enumerate(candidates):
        row = _candidate_to_dict(c)
        output_rows[i] = row

    # =========================================================================
    # Stage 1: GATE
    # =========================================================================
    # Gate order (RFC §4.4): SCHEMA_ERROR → DISABLED → LOW_N → NEG_EV_NET
    # First matching reason wins; survivors get status=None
    for i, c in enumerate(candidates):
        row = output_rows[i]

        # Check SCHEMA_ERROR first (earliest in gate order)
        if validate_candidate(c) is not None:
            row["status"] = "SCHEMA_ERROR"
            row["flags"] = []
            row["derived"] = {}  # No derived for invalid schema
            continue

        # Compute derived fields (needed for ev_net and later stages)
        derived = compute_derived(c, config)
        row["derived"] = derived

        # Compute ev_ratio and DATA_MISMATCH flag
        ev_ratio, is_mismatch = compute_ev_ratio(c, derived)
        row["flags"] = ["DATA_MISMATCH"] if is_mismatch else []

        # Check DISABLED (second in gate order)
        if not enabled_mask.get(c.ticker, True):
            row["status"] = "DISABLED"
            continue

        # Check LOW_N (third in gate order)
        if c.n < config.min_n:
            row["status"] = "LOW_N"
            continue

        # Check NEG_EV_NET (fourth in gate order)
        if derived["ev_net"] <= 0:
            row["status"] = "NEG_EV_NET"
            continue

        # Survived gating
        row["status"] = None

    # =========================================================================
    # Stage 2: PER-ASSET COLLAPSE (on survivors only)
    # =========================================================================
    # Group survivors by ticker; check for direction conflict or mark duplicates
    ticker_groups = defaultdict(list)

    for i, c in enumerate(candidates):
        row = output_rows[i]
        if row["status"] is not None:
            continue  # Skip rejected rows
        ticker_groups[c.ticker].append((i, c, row))

    # For each ticker, check direction conflict or mark duplicates
    for ticker, group in ticker_groups.items():
        directions = set(c.direction for _, c, _ in group)

        if len(directions) > 1:
            # Both long and short survived gating for this ticker
            # Reject all rows for that ticker with DIRECTION_CONFLICT
            for _, _, row in group:
                row["status"] = "DIRECTION_CONFLICT"
        else:
            # Only one direction; keep max-score row, reject rest as DUP_ASSET
            if len(group) > 1:
                # Find index of max-score row
                max_idx = max(
                    range(len(group)),
                    key=lambda j: group[j][2]["derived"]["score"]
                )
                for j, (_, _, row) in enumerate(group):
                    if j != max_idx:
                        row["status"] = "DUP_ASSET"

    # =========================================================================
    # Stage 3: RANK + TOP-K (on survivors only)
    # =========================================================================
    # Collect survivors after collapse, maintaining original index order
    survivors = []
    for i, c in enumerate(candidates):
        row = output_rows[i]
        if row["status"] is None:
            survivors.append((i, row, c))

    # Sort by score descending (stable sort preserves insertion order on ties)
    # per RFC §4.4's deterministic tie-break note
    survivors.sort(key=lambda x: x[1]["derived"]["score"], reverse=True)

    # Mark top-K as selected (status remains None), rest as BELOW_TOPK
    for rank, (_, row, _) in enumerate(survivors):
        if rank >= config.top_k:
            row["status"] = "BELOW_TOPK"

    # Return all rows in original input order (not sorted)
    return [output_rows[i] for i in range(len(candidates))]


def _candidate_to_dict(c: Candidate) -> dict:
    """Convert Candidate dataclass to dict with all fields.

    Returns a dict ready for output, with all Candidate fields plus placeholders
    for derived, status, and flags fields to be filled in by select_candidates().
    """
    return {
        "strategy": c.strategy,
        "ticker": c.ticker,
        "direction": c.direction,
        "entry": c.entry,
        "stop": c.stop,
        "target": c.target,
        "ev_pct": c.ev_pct,
        "base_win_rate": c.base_win_rate,
        "n": c.n,
        "backtest_period": c.backtest_period,
        "sharpe": c.sharpe,
        "advised_liquidity_pct": c.advised_liquidity_pct,
        "avg_win_pct": c.avg_win_pct,
        "avg_loss_pct": c.avg_loss_pct,
        "avg_holding_days": c.avg_holding_days,
        "model": c.model,
        "derived": {},  # Filled in by select_candidates
        "status": None,  # Filled in by select_candidates
        "flags": [],  # Filled in by select_candidates
    }


def size_selected(survivors: list[dict], config: AllocationConfig) -> list[dict]:
    """Apply position cap, cluster caps, gross cap, and dust filter to top-K survivors.

    Implements RFC §4.4 sizing pipeline (position cap → cluster caps → gross cap → dust filter)
    on the survivors from select_candidates(). Returns the full row set (all candidates, both
    selected and rejected) with final sizing applied to rows with status=None (top-K).

    Args:
        survivors: Full output list from select_candidates(), containing both rejected rows
                  (with status set) and top-K survivors (with status=None)
        config: AllocationConfig with sizing parameters (max_pos_pct, max_cluster_pct,
               gross_cap_pct, dust_min_pct, cluster_map)

    Returns:
        List of dicts with the same rows as input, each top-K survivor now having:
        - alloc: final allocation % after all caps and dust filter
        - status: "SELECTED" or "DUST" (for survivors); unchanged for rejected rows
        - flags: may include "POS_CAPPED" and/or "CLUSTER_CAPPED" (for survivors)
        Rejected rows pass through unchanged.

    Per RFC §4.6, dust filter is single-pass: zeroed allocations are not redistributed.
    """
    from collections import defaultdict

    # Deep copy to avoid mutating input
    result = [dict(row) for row in survivors]

    # =========================================================================
    # Stage 1: POSITION CAP (per-row, on survivors only)
    # =========================================================================
    # alloc = min(kelly_frac * 100, max_pos_pct)
    for row in result:
        if row["status"] is not None:
            # Rejected row; skip sizing
            continue

        kelly_frac = row["derived"]["kelly_frac"]
        alloc_raw = min(kelly_frac * 100, config.max_pos_pct)

        # Flag if capped
        if kelly_frac * 100 > config.max_pos_pct:
            if "POS_CAPPED" not in row["flags"]:
                row["flags"].append("POS_CAPPED")

        row["alloc"] = alloc_raw

    # =========================================================================
    # Stage 2: CLUSTER CAPS (proportional scale-down within over-cap clusters)
    # =========================================================================
    # Group post-cap allocations by cluster; if cluster sum > max_cluster_pct,
    # scale that cluster's allocations proportionally
    cluster_groups = defaultdict(list)
    for row in result:
        if row["status"] is not None:
            # Rejected row; skip
            continue

        ticker = row["ticker"]
        # Unmapped tickers form their own singleton cluster (their own ticker
        # name as cluster key), so unrelated unmapped tickers never get
        # cluster-capped together under one shared bucket.
        cluster = config.cluster_map.get(ticker, ticker)
        cluster_groups[cluster].append(row)

    for cluster, cluster_rows in cluster_groups.items():
        cluster_sum = sum(row["alloc"] for row in cluster_rows)

        if cluster_sum > config.max_cluster_pct:
            # Scale factor: new_sum / old_sum
            scale_factor = config.max_cluster_pct / cluster_sum

            for row in cluster_rows:
                row["alloc"] *= scale_factor

                # Add CLUSTER_CAPPED flag if not already present
                if "CLUSTER_CAPPED" not in row["flags"]:
                    row["flags"].append("CLUSTER_CAPPED")

    # =========================================================================
    # Stage 3: GROSS CAP (proportional scale-down if total > gross_cap_pct)
    # =========================================================================
    survivors_with_status_none = [row for row in result if row["status"] is None]
    gross_sum = sum(row["alloc"] for row in survivors_with_status_none)

    if gross_sum > config.gross_cap_pct:
        scale_factor = config.gross_cap_pct / gross_sum

        for row in survivors_with_status_none:
            row["alloc"] *= scale_factor

    # =========================================================================
    # Stage 4: DUST FILTER (single-pass, no redistribution)
    # =========================================================================
    # Any row with final alloc < dust_min_pct gets alloc=0 and status="DUST"
    for row in result:
        if row["status"] is not None:
            # Rejected row; skip
            continue

        if row["alloc"] < config.dust_min_pct:
            row["alloc"] = 0.0
            row["status"] = "DUST"
        else:
            row["status"] = "SELECTED"

    return result


def allocate(candidates: list[Candidate], config: AllocationConfig, enabled_mask: dict = None) -> AllocationResult:
    """Top-level allocation orchestration per RFC allocation_sheet.md §8.

    Implements the complete selection and sizing pipeline:
    1. validate_candidate (E11-S02): schema validation
    2. compute_derived (E11-S03): per-row derived metrics
    3. compute_ev_ratio (E11-S04): data quality check
    4. select_candidates (E11-S05): gating, collapse, top-K ranking
    5. size_selected (E11-S06): position/cluster/gross caps + dust filter

    This is the reference oracle per RFC §4 and §8 ("pure function of (candidates, config)").

    Args:
        candidates: list of Candidate objects
        config: AllocationConfig with all sizing/gating parameters
        enabled_mask: dict mapping ticker -> bool; defaults to all-enabled (empty dict
                     since select_candidates defaults missing tickers to enabled).
                     Pass None to use default.

    Returns:
        AllocationResult with:
        - rows: full output from size_selected (all candidates, both selected and rejected)
        - selected_count: number of rows with status == "SELECTED"
        - gross_exposure_pct: sum of alloc for status == "SELECTED" rows
        - rejection_counts: dict mapping rejection status -> count for all non-selected rows
    """
    if enabled_mask is None:
        enabled_mask = {}

    # Run the full pipeline: validate -> derive -> collapse -> size
    result_rows = select_candidates(candidates, config, enabled_mask)
    result_rows = size_selected(result_rows, config)

    # Compute summary statistics
    selected_rows = [row for row in result_rows if row.get("status") == "SELECTED"]
    selected_count = len(selected_rows)

    gross_exposure_pct = sum(row.get("alloc", 0.0) for row in selected_rows)

    # Rejection counts: all non-SELECTED rows
    rejection_counts = {}
    for row in result_rows:
        status = row.get("status")
        if status != "SELECTED":
            rejection_counts[status] = rejection_counts.get(status, 0) + 1

    return AllocationResult(
        rows=result_rows,
        selected_count=selected_count,
        gross_exposure_pct=gross_exposure_pct,
        rejection_counts=rejection_counts,
    )


def load_cluster_map(path: str) -> dict:
    """Load ticker-to-cluster mapping from a CSV file.

    Reads a two-column CSV file (no header) with format:
        ticker,cluster_name

    Missing tickers in the map should NOT crash size_selected(). The clustering
    logic in size_selected() implements a fallback: unmapped tickers form their
    own singleton cluster (using ticker name as cluster key).

    Args:
        path: file path to CSV file

    Returns:
        dict[str, str] mapping ticker -> cluster_name

    Raises:
        FileNotFoundError: if path does not exist
        ValueError: if CSV parsing fails (e.g., missing columns)
    """
    cluster_map = {}

    with open(path, 'r') as f:
        reader = csv.reader(f)
        for row_num, row in enumerate(reader, start=1):
            if len(row) < 2:
                raise ValueError(f"Line {row_num} has fewer than 2 columns")
            ticker = row[0].strip()
            cluster_name = row[1].strip()
            if ticker:  # Skip empty ticker lines
                cluster_map[ticker] = cluster_name

    return cluster_map


def _is_missing(value):
    """Check if a value is missing (None or NaN).

    Replicates the helper from kairos_signals.py for consistency.
    """
    import numpy as np

    if value is None:
        return True
    try:
        return bool(np.isnan(value))
    except (TypeError, ValueError):
        return False


# ===========================================================================
# E11-S08: Formula Template Engine (XLSX/ODS dialect rendering)
# ===========================================================================

"""Formula templates for portfolio allocation computations.

Each formula template is written once in a canonical XLSX form with:
- Cell references using Excel A1 notation (e.g., F20, G$3)
- Config cell references with absolute row anchors (e.g., $E$3, $E$4)
- Placeholder tokens for row substitution: {row}
- Function calls using comma argument separators (XLSX style)

Templates are rendered per dialect via render_formula(name, row, fmt):
- XLSX: Returns formula with '=' prefix and comma separators
- ODS: Returns formula with 'of:=' prefix and semicolon separators
  (per ODS spec; semicolons separate function arguments in many locales)

Column mapping (formula columns P through AN, per E11-S08/S09/S10/S15 + the
Enabled-column/redistribution update):
- P  = ev_net              EV net of costs (RFC §4.2), fraction
- Q  = ev_total            EV as % of total equity = ev_net_raw_pct * alloc_pct (both fractions)
- R  = kelly_raw           Binary Kelly before capping/multiplier
- S  = score               Return per unit risk (ranking key)
- T  = alloc_pct           Live-redistributed allocation fraction among enabled rows (see column A)
- U  = alloc_eur           Allocation expressed as currency amount
- V  = flags               Composite flags (DATA_MISMATCH, POS_CAPPED, CLUSTER_CAPPED)
- W  = advised_liq         Carried upstream liquidity % (displayed but ignored)
- X  = risk_pct            ABS(stop - entry) / entry, fraction
- Y  = reward_pct          ABS(target - entry) / entry, fraction
- Z  = b                   Payoff ratio (geometry fallback in v1)
- AA = loss_pct            Basis for score denominator
- AB = shrink              Confidence weight n / (n + n0)
- AC = ev_shrunk           ev_pct * shrink
- AD = p_shrunk            Win rate shrunk toward 50%
- AE = kelly_frac          Fractional Kelly decimal (before cap)
- AF = alloc_raw_pct       kelly_frac, fraction (no *100)
- AG = pos_capped_alloc    MIN(alloc_raw_pct, max_pos_pct)
- AH = pos_capped_flag     "POS_CAPPED" if capped
- AI = ev_implied          base_win_rate*reward - (1-base_win_rate)*risk
- AJ = ev_ratio            ev_pct / ev_implied
- AK = data_mismatch       "DATA_MISMATCH" if ev_ratio outside [0.5, 2.0]
- AL = cluster_scale       Cluster-level scale factor (SUMIFS over pos_capped_alloc)
- AM = enabled_flag        1 if column A is blank or "true" (case-insensitive), else 0
- AN = base_alloc_pct      Pre-redistribution allocation (old alloc_pct formula: AG*AL*gross_scale)

Column A (Enabled) is the new leftmost, user-editable plain-text cell per
data row: "true" (case-insensitive, trimmed) or blank means enabled;
anything else means disabled. Disabled rows drop out of the live
redistribution in column T entirely; the remaining enabled rows' Alloc %
(column T) is renormalized to sum to 100% based on their AN (base_alloc_pct)
weights. This is a live spreadsheet formula (SUMPRODUCT over AN/AM), not a
Python-side computation.

All percent-labeled quantities are stored as 0-1 fractions in the sheet and
displayed via percent number format; only the Python domain (Candidate,
AllocationConfig, allocate()) keeps the 0-100 scale.

Summary-block config cell layout (absolute references):
- $E$3: n0 (shrinkage constant)
- $E$4: round_trip_cost_pct (cost per round trip, fraction)
- $E$5: kelly_mult (fractional Kelly multiplier)
- $E$6: gross_cap_pct (total exposure cap, fraction)
- $E$7: max_pos_pct (per-position cap, fraction)
- $E$8: max_cluster_pct (per-cluster cap, fraction)
- $E$9: equity (optional account equity for currency column; blank = no amount)
- $E$14: gross_scale factor summary cell (proportional scale if over gross cap)

Data cell layout for row N (per RFC §5.2 static columns A-O):
- A: Enabled ("true"/blank = enabled, anything else = disabled)
- C: Cluster
- F: Entry price
- G: Stop loss price
- H: Target price
- L: n (number of backtest trades)
- M: Win raw (base_win_rate as a fraction, e.g., 0.47)
- O: EV raw % (ev_pct from backtest)
"""

# Template strings keyed by column letter (O..AK) plus the summary factor.
# Each contains a single {row} placeholder substituted by render_formula().
# All percent-like quantities are fractions (0-1); no *100/ /100 conversions.
_FORMULA_TEMPLATES = {
    # Visible Section A derived columns
    "P":  "O{row}*AB{row}-$E$4",                      # ev_net
    "Q":  "O{row}*T{row}",                            # ev_total = EV raw % * Alloc % (now the redistributed alloc)
    "R":  "IFERROR(AD{row}-(1-AD{row})/Z{row},0)",    # kelly_raw
    "S":  "IFERROR(P{row}/AA{row},0)",                # score
    "T":  "IF(AM{row}=0,0,IFERROR(AN{row}/SUMPRODUCT(AN$21:AN$401,AM$21:AM$401),0))",  # alloc_pct: redistribute base_alloc among enabled rows, renormalized to 100%
    "U":  'IF($E$9="","",T{row}*$E$9)',               # alloc_eur (uses the new, final Alloc %)
    "V":  'IF(AK{row}="","",AK{row})&IF(AH{row}="","",IF(AK{row}="",""," ")&AH{row})&IF(AL{row}<1,IF(AND(AK{row}="",AH{row}=""),""," ")&"CLUSTER_CAPPED","")',  # flags
    "W":  '""',                                        # advised_liq (ignored)

    # Helper columns (grouped/collapsed by the sheet writer)
    "X": "IF(F{row}=0,0,ABS(G{row}-F{row})/F{row})",  # risk_pct
    "Y": "IF(F{row}=0,0,ABS(H{row}-F{row})/F{row})",  # reward_pct
    "Z": "IF(AA{row}=0,0,Y{row}/AA{row})",            # b
    "AA": "IF(F{row}=0,0,X{row})",                     # loss_pct
    "AB": "IF(L{row}=0,0,L{row}/(L{row}+$E$3))",       # shrink
    "AC": "O{row}*AB{row}",                            # ev_shrunk
    "AD": "0.5+(M{row}-0.5)*AB{row}",                  # p_shrunk
    "AE": "MAX(R{row},0)*$E$5",                        # kelly_frac
    "AF": "AE{row}",                                    # alloc_raw_pct
    "AG": "MIN(AF{row},$E$7)",                          # pos_capped_alloc
    "AH": 'IF(AF{row}>$E$7,"POS_CAPPED","")',           # pos_capped_flag
    "AI": "M{row}*Y{row}-(1-M{row})*X{row}",            # ev_implied
    "AJ": "IFERROR(O{row}/AI{row},0)",                  # ev_ratio
    "AK": 'IF(OR(AJ{row}<0.5,AJ{row}>2),"DATA_MISMATCH","")',  # data_mismatch
    "AL": "IF(SUMIFS(AG$21:AG$401,C$21:C$401,C{row})>$E$8,$E$8/SUMIFS(AG$21:AG$401,C$21:C$401,C{row}),1)",  # cluster_scale
    "AM": 'IF(OR(A{row}="",LOWER(TRIM(A{row}))="true"),1,0)',  # enabled_flag
    "AN": "AG{row}*AL{row}*$E$14",                      # base_alloc_pct (the old alloc_pct formula, relocated)

    # Summary-block gross scale factor (rendered into $E$14).
    # Must scale the post-cluster-cap total (AG * AL) down to gross_cap_pct.
    # SUM(AL) alone is wrong: it is just the count of cluster-scale factors.
    "gross_scale": "IF(SUMPRODUCT(AG$21:AG$401,AL$21:AL$401)>$E$6,$E$6/SUMPRODUCT(AG$21:AG$401,AL$21:AL$401),1)",
}


# Concept-name aliases for callers that prefer readable names.  Each alias
# resolves to one of the canonical column-letter keys above, so the same
# single template is used for both naming styles.
_FORMULA_ALIASES = {
    "ev_net": "P",
    "ev_total": "Q",
    "kelly_raw": "R",
    "score": "S",
    "alloc_pct": "T",
    "alloc_eur": "U",
    "flags": "V",
    "advised_liq_pct": "W",
    "risk_pct": "X",
    "reward_pct": "Y",
    "b": "Z",
    "loss_pct": "AA",
    "shrink": "AB",
    "ev_shrunk": "AC",
    "p_shrunk": "AD",
    "kelly_frac": "AE",
    "alloc_raw_pct": "AF",
    "pos_capped_alloc": "AG",
    "pos_capped": "AH",
    "ev_implied": "AI",
    "ev_ratio": "AJ",
    "data_mismatch": "AK",
    "cluster_scale": "AL",
    "enabled_flag": "AM",
    "base_alloc_pct": "AN",
    "gross_scale": "gross_scale",
}


def render_formula(name: str, row: int, fmt: str) -> str:
    """Render a formula template for a given row number and dialect.

    Implements E11-S08 acceptance criteria: both XLSX and ODS dialects
    derive from one shared template, with only dialect-specific syntax changes.

    Args:
        name: Formula name.  Accepts canonical column-letter keys ("P".."AN",
            "gross_scale") and concept aliases ("risk_pct", "ev_net", ...).
        row: Data row number (21..401 per RFC §5.1), used for cell reference substitution.
        fmt: Output dialect, "xlsx" or "ods".

    Returns:
        Formula string ready for insertion into a spreadsheet cell:
        - XLSX: "=<formula>" with comma-separated function arguments
        - ODS: "of:=<formula>" with semicolon-separated function arguments

    Raises:
        ValueError: if name cannot be resolved or fmt not in ("xlsx", "ods")

    Notes:
        - Row-number substitution is exact: {row} placeholders become the literal row number
        - Config cell references like $E$3 are preserved as-is
        - Both dialects use the same underlying template; no separate formula sets
        - Compliance: forbids FILTER, SORT, UNIQUE, LET, LAMBDA, XLOOKUP, MAXIFS, MINIFS
    """
    if fmt not in ("xlsx", "ods"):
        raise ValueError(f"fmt must be 'xlsx' or 'ods', got {fmt!r}")

    canonical = _FORMULA_ALIASES.get(name, name)
    if canonical not in _FORMULA_TEMPLATES:
        raise ValueError(f"formula name {name!r} not found in templates")

    template = _FORMULA_TEMPLATES[canonical]
    formula = template.format(row=row)

    if fmt == "xlsx":
        return "=" + formula
    return "of:=" + _convert_commas_to_semicolons(formula)


def _convert_commas_to_semicolons(formula: str) -> str:
    """Convert comma argument separators to semicolons for ODS format.

    The canonical templates use commas only as function-argument separators and
    inside decimal numbers (e.g. 0.5) use a dot, so a global replacement is safe.
    No string literal contains a comma.
    """
    return formula.replace(",", ";")


def get_formula_names() -> list[str]:
    """Return the sorted list of canonical formula names (column letters + gross_scale)."""
    return sorted(_FORMULA_TEMPLATES.keys())


def get_formula_aliases() -> dict[str, str]:
    """Return the concept-name -> canonical-key alias map."""
    return dict(_FORMULA_ALIASES)


# =============================================================================
# E11-S09 / E11-S10: Shared Sheet Layout Constants
# =============================================================================

# Config block layout.  Order is chosen so that the first seven editable value
# cells ($E$3..$E$9) line up with the absolute references used by the formula
# templates above.
_CONFIG_BLOCK = [
    ("n0", "n0"),
    ("round_trip_cost_pct", "round_trip_cost_pct"),
    ("kelly_mult", "kelly_mult"),
    ("gross_cap_pct", "gross_cap_pct"),
    ("max_pos_pct", "max_pos_pct"),
    ("max_cluster_pct", "max_cluster_pct"),
    ("equity", "equity"),
    ("min_n", "min_n"),
    ("top_k", "top_k"),
    ("dust_min_pct", "dust_min_pct"),
    ("cluster_map", "cluster_map"),
]

_DATA_START_ROW = 20
_HEADER_ROW = 20

# Static columns A-O and their human-readable header. Column A (Enabled) is
# the new leftmost user-editable plain-text enable/disable cell.
_STATIC_HEADERS = [
    "Enabled", "Ticker", "Cluster", "Strategy", "Dir", "Entry", "Stop", "Target",
    "Risk %", "Reward %", "b", "n", "Win raw", "Win shrunk", "EV raw %",
]

# Headers for the formula-driven columns P-W (visible Section A derived columns).
_FORMULA_HEADERS = {
    "P": "EV net %",
    "Q": "EV total",
    "R": "Kelly raw",
    "S": "Score",
    "T": "Alloc %",
    "U": "Alloc EUR",
    "V": "Flags",
    "W": "Advised liq % (ignored)",
}

# Headers for the helper formula columns X-AN.
_HELPER_HEADERS = [
    "risk_pct", "reward_pct", "b", "loss_pct", "shrink", "ev_shrunk",
    "p_shrunk", "kelly_frac", "alloc_raw_pct", "pos_capped_alloc",
    "pos_capped_flag", "ev_implied", "ev_ratio", "data_mismatch", "cluster_scale",
    "enabled_flag", "base_alloc_pct",
]

_FORMULA_COLS = [
    "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA",
    "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN",
]

# Trailing plain-value Model column, appended one position to the right of the
# formula/helper block (AN). This is NOT part of _FORMULA_COLS: it carries a
# plain string value (no formula) and must never shift the position-sensitive
# A-AN layout (see module docstring on column AM/AN redistribution formulas).
_MODEL_COL = "AO"

# Column letters (percent-labeled) that receive "0.0%" number format / ODS
# percent cell style: visible Risk %, Reward %, EV raw %, EV net %, EV total,
# Alloc %, plus helper risk_pct, reward_pct, loss_pct, alloc_raw_pct, and the
# new base_alloc_pct helper (same kind of quantity as alloc_pct).
_PERCENT_COLS = ("I", "J", "O", "P", "Q", "T", "X", "Y", "AA", "AF", "AN")

# Summary-block formulas shared by XLSX and ODS writers.  Dialect-specific
# rendering (comma vs. semicolon separators, '=' vs. 'of:=' prefix) is applied
# by the writers so the formula logic is not duplicated.
_SUMMARY_FORMULAS = {
    "selected_count": '=COUNTIF(T21:T401,">0")',
    "gross_exposure": "=SUM(T21:T401)",
    "enabled_count": "=SUM(AM21:AM401)",
    "ev_total_sum": "=SUM(Q21:Q401)",
}


def _pct_to_frac(value):
    """Convert a 0-100 percent value to a 0-1 fraction. None-safe."""
    if value is None:
        return None
    return value / 100.0


def _sorted_for_sheet(rows):
    """Sort rows for sheet display: selected/allocated rows by EV total desc.

    alloc > 0 rows come first (sorted by ev_total descending); everything
    else (rejected/zero-alloc rows) keeps its original relative order at the
    bottom. Uses a stable sort so ties preserve input order.
    """
    def key(r):
        alloc = r.get("alloc") or 0.0
        ev_pct = (r.get("ev_pct") or 0.0) / 100.0
        ev_total = ev_pct * alloc / 100.0
        return (0 if alloc > 0 else 1, -ev_total)

    return sorted(rows, key=key)


def _xlsx_column_letter_to_index(col_letter: str) -> int:
    """Convert an Excel column letter (A, AA, AJ) to a 1-based column index."""
    idx = 0
    for ch in col_letter:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx


def _empty_if_none(value):
    """Return an empty string for None, otherwise the value unchanged."""
    if value is None:
        return ""
    return value


def _xlsx_formula_to_ods(formula: str) -> str:
    """Convert an XLSX-style formula string into an ODS formula string.

    Applies the same comma-to-semicolon conversion used by render_formula()
    for ODS output and prefixes the result with ``of:=``.
    """
    if not formula.startswith("="):
        return formula
    return "of:=" + _convert_commas_to_semicolons(formula[1:])


# =============================================================================
# E11-S09: XLSX Sheet Writer
# =============================================================================


def write_xlsx_sheet(workbook, result: AllocationResult, config: AllocationConfig,
                     report_date, generator_version: str):
    """Write the ``Allocation`` sheet into an existing ``openpyxl.Workbook``.

    Layout follows RFC allocation_sheet.md §5 and ticket E11-S09/S15 (plus the
    Enabled-column/live-redistribution update):
      - Row 1: title, report date, generator version.
      - Rows 3-13: config block (parameter name, editable value, locked default).
      - Rows 14-17: summary formulas (selected count, gross exposure %, enabled
        count, EV total, and the gross scale factor in $E$14).
      - Row 18: blank spacer row.
      - Row 19: instruction line.
      - Row 20: header row A-AO (Model in AO is a trailing plain value).
      - Rows 21..N: one row per candidate, sorted by EV total descending
        (``_sorted_for_sheet``).
      - Below the data: Section B cluster-exposure table.
      - Helper columns X-AN are grouped (outlined) but not hidden.
      - All cells are locked except the editable config value cells (E3:E13),
        column A (the Enabled per-row toggle) and column O (the per-row
        editable EV raw % input column).
      - Percent-labeled cells are 0-1 fractions with "0.0%" number format.
      - Autofilter is enabled on the header/data range.

    The function performs no disk I/O; it mutates the provided in-memory
    ``Workbook``.
    """
    from openpyxl import Workbook

    if not isinstance(workbook, Workbook):
        raise TypeError("workbook must be an openpyxl Workbook")

    ws = workbook.create_sheet("Allocation")

    # -------------------------------------------------------------------------
    # Row 1: title line
    # -------------------------------------------------------------------------
    ws["A1"] = "Portfolio Allocation"
    ws["B1"] = _empty_if_none(report_date)
    ws["C1"] = f"generator {generator_version}"

    # -------------------------------------------------------------------------
    # Rows 3-13: config block
    # -------------------------------------------------------------------------
    # Parameter names in column D, editable values in column E, shipped defaults
    # (locked) in column F.
    default_config = AllocationConfig()
    pct_value_cells = []  # (row, col) cells holding a pct fraction, for number_format
    for offset, (label, attr) in enumerate(_CONFIG_BLOCK):
        row = 3 + offset
        ws.cell(row=row, column=4, value=label)
        is_pct = "pct" in label
        value = getattr(config, attr)
        if attr == "cluster_map":
            value = str(value) if value else ""
        elif is_pct:
            value = _pct_to_frac(value)
        ws.cell(row=row, column=5, value=_empty_if_none(value))
        default_value = getattr(default_config, attr)
        if attr == "cluster_map":
            default_value = str(default_value) if default_value else ""
        elif is_pct:
            default_value = _pct_to_frac(default_value)
        ws.cell(row=row, column=6, value=_empty_if_none(default_value))
        if is_pct:
            pct_value_cells.append((row, 5))
            pct_value_cells.append((row, 6))

    # -------------------------------------------------------------------------
    # Rows 14-17: summary block
    # -------------------------------------------------------------------------
    ws["A14"] = "Selected count"
    ws["D14"] = _SUMMARY_FORMULAS["selected_count"]
    ws["A15"] = "Gross exposure %"
    ws["D15"] = _SUMMARY_FORMULAS["gross_exposure"]
    ws["A16"] = "Enabled count"
    ws["D16"] = _SUMMARY_FORMULAS["enabled_count"]
    ws["A17"] = "EV total"
    ws["D17"] = _SUMMARY_FORMULAS["ev_total_sum"]
    ws["E14"] = render_formula("gross_scale", 14, "xlsx")

    # -------------------------------------------------------------------------
    # Row 18: blank spacer row (intentionally left empty)
    # -------------------------------------------------------------------------

    # -------------------------------------------------------------------------
    # Row 19: instruction line
    # -------------------------------------------------------------------------
    ws["A19"] = (
        "Edit only the config values (column E), the per-row input column "
        "(column O), and the Enabled column (column A). All other cells are computed."
    )

    # -------------------------------------------------------------------------
    # Row 20: header row
    # -------------------------------------------------------------------------
    headers = (
        _STATIC_HEADERS
        + [_FORMULA_HEADERS[col] for col in _FORMULA_COLS[:8]]
        + _HELPER_HEADERS
        + ["Model"]
    )
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=_HEADER_ROW, column=col_idx, value=header)

    # -------------------------------------------------------------------------
    # Rows 21..N: candidate rows, sorted by EV total descending
    # -------------------------------------------------------------------------
    def _cluster_for_ticker(ticker: str) -> str:
        return config.cluster_map.get(ticker, ticker)

    sorted_rows = _sorted_for_sheet(result.rows)

    for row_offset, row_data in enumerate(sorted_rows):
        excel_row = _DATA_START_ROW + 1 + row_offset
        derived = row_data.get("derived", {}) or {}
        ticker = row_data.get("ticker", "")
        direction = row_data.get("direction", "")
        enabled_value = "true" if row_data.get("status") == "SELECTED" else "false"

        static_values = [
            enabled_value,
            ticker,
            _cluster_for_ticker(ticker),
            row_data.get("strategy", ""),
            direction.capitalize() if isinstance(direction, str) else "",
            row_data.get("entry"),
            row_data.get("stop"),
            row_data.get("target"),
            _pct_to_frac(derived.get("risk_pct")),
            _pct_to_frac(derived.get("reward_pct")),
            derived.get("b"),
            row_data.get("n"),
            row_data.get("base_win_rate"),
            derived.get("p_shrunk"),
            _pct_to_frac(row_data.get("ev_pct")),
        ]
        for col_idx, value in enumerate(static_values, start=1):
            ws.cell(row=excel_row, column=col_idx, value=_empty_if_none(value))

        for col_letter in _FORMULA_COLS:
            col_idx = _xlsx_column_letter_to_index(col_letter)
            ws.cell(row=excel_row, column=col_idx,
                    value=render_formula(col_letter, excel_row, "xlsx"))

        # Trailing plain-value Model column, appended after all A-AN columns.
        model_col_idx = _xlsx_column_letter_to_index(_MODEL_COL)
        ws.cell(row=excel_row, column=model_col_idx,
                value=row_data.get("model") or "")

    # -------------------------------------------------------------------------
    # Percent number formatting: visible + helper percent-labeled columns,
    # over the data rows, plus config/summary cells noted above.
    # -------------------------------------------------------------------------
    data_end_row = _DATA_START_ROW + len(sorted_rows)
    for col_letter in _PERCENT_COLS:
        col_idx = _xlsx_column_letter_to_index(col_letter)
        for row in range(_DATA_START_ROW + 1, data_end_row + 1):
            ws.cell(row=row, column=col_idx).number_format = "0.0%"
    for row, col in pct_value_cells:
        ws.cell(row=row, column=col).number_format = "0.0%"
    ws["D15"].number_format = "0.0%"
    ws["D17"].number_format = "0.0%"

    # -------------------------------------------------------------------------
    # Autofilter over header + data rows
    # -------------------------------------------------------------------------
    ws.auto_filter.ref = f"A{_HEADER_ROW}:{_MODEL_COL}{data_end_row}"

    # -------------------------------------------------------------------------
    # Section B: cluster exposure table
    # -------------------------------------------------------------------------
    cluster_header_row = data_end_row + 2
    ws.cell(row=cluster_header_row, column=1, value="Cluster")
    ws.cell(row=cluster_header_row, column=2, value="Positions")
    ws.cell(row=cluster_header_row, column=3, value="Gross %")
    ws.cell(row=cluster_header_row, column=4, value="Cap %")
    ws.cell(row=cluster_header_row, column=5, value="Capped?")

    selected_rows = [r for r in result.rows if r.get("status") == "SELECTED"]
    clusters = sorted(set(config.cluster_map.values())) if config.cluster_map else []
    for cluster_offset, cluster in enumerate(clusters):
        row = cluster_header_row + 1 + cluster_offset
        cluster_rows = [
            r for r in selected_rows
            if _cluster_for_ticker(r.get("ticker", "")) == cluster
        ]
        positions = len(cluster_rows)
        gross = sum(r.get("alloc", 0.0) for r in cluster_rows)
        capped = any("CLUSTER_CAPPED" in r.get("flags", []) for r in cluster_rows)
        ws.cell(row=row, column=1, value=cluster)
        ws.cell(row=row, column=2, value=positions)
        ws.cell(row=row, column=3, value=_pct_to_frac(gross)).number_format = "0.0%"
        ws.cell(row=row, column=4, value=_pct_to_frac(config.max_cluster_pct)).number_format = "0.0%"
        ws.cell(row=row, column=5, value="yes" if capped else "no")

    # -------------------------------------------------------------------------
    # Section C: rejected signals (compact audit trail)
    # -------------------------------------------------------------------------
    if cluster_header_row + len(clusters) + 1 >= ws.max_row:
        rejected_header_row = cluster_header_row + len(clusters) + 2
    else:
        rejected_header_row = ws.max_row + 2

    ws.cell(row=rejected_header_row, column=1, value="Ticker")
    ws.cell(row=rejected_header_row, column=2, value="Strategy")
    ws.cell(row=rejected_header_row, column=3, value="Dir")
    ws.cell(row=rejected_header_row, column=4, value="Score")
    ws.cell(row=rejected_header_row, column=5, value="Reason")

    rejected = [r for r in result.rows if r.get("status") != "SELECTED"]
    rejected.sort(
        key=lambda r: (
            r.get("status", ""),
            -(r.get("derived", {}) or {}).get("score", float("-inf")),
        )
    )
    for rej_offset, r in enumerate(rejected):
        row = rejected_header_row + 1 + rej_offset
        direction = r.get("direction", "")
        score = (r.get("derived", {}) or {}).get("score")
        ws.cell(row=row, column=1, value=r.get("ticker", ""))
        ws.cell(row=row, column=2, value=r.get("strategy", ""))
        ws.cell(row=row, column=3,
                value=direction.capitalize() if isinstance(direction, str) else "")
        ws.cell(row=row, column=4, value=_empty_if_none(score))
        ws.cell(row=row, column=5, value=r.get("status", ""))

    # -------------------------------------------------------------------------
    # Column grouping: helper columns X-AN are outlined but not hidden.
    # -------------------------------------------------------------------------
    ws.column_dimensions.group("X", "AN", outline_level=1, hidden=False)

    # -------------------------------------------------------------------------
    # Sheet protection: lock everything except the editable config values
    # (E3:E13), column A (Enabled), and column O (EV raw % input).
    # -------------------------------------------------------------------------
    for cell in ws["A"]:
        cell.protection = Protection(locked=False)

    for cell in ws["O"]:
        cell.protection = Protection(locked=False)

    for config_row in range(3, 3 + len(_CONFIG_BLOCK)):
        ws.cell(row=config_row, column=5).protection = Protection(locked=False)

    ws.protection.sheet = True


# =============================================================================
# E11-S10: ODS Sheet Writer
# =============================================================================

_ODS_PERCENT_STYLE_NAME = "KairosPercent"
_ODS_PERCENT_CELL_STYLE_NAME = "KairosPercentCell"


def _ods_ensure_percent_style(document):
    """Ensure the document has a percent number style + cell style defined.

    Idempotent: safe to call multiple times for the same document (odfpy does
    not de-duplicate automaticstyles, so callers should only call this once
    per document, but re-adding is otherwise harmless).
    """
    from odf.number import Number, PercentageStyle, Text
    from odf.style import Style

    percent_style = PercentageStyle(name=_ODS_PERCENT_STYLE_NAME)
    percent_style.addElement(Number(decimalplaces="1", minintegerdigits="1"))
    percent_style.addElement(Text(text="%"))
    document.automaticstyles.addElement(percent_style)

    cell_style = Style(
        name=_ODS_PERCENT_CELL_STYLE_NAME,
        family="table-cell",
        datastylename=_ODS_PERCENT_STYLE_NAME,
    )
    document.automaticstyles.addElement(cell_style)


def _ods_create_cell(value, formula=None, percent=False):
    """Create an odfpy ``TableCell`` with optional ODS formula.

    Args:
        percent: if True, apply the percent cell style (``KairosPercentCell``)
            so the numeric value renders as ``0.0%``.
    """
    from odf.table import TableCell
    from odf.text import P

    text = "" if value is None else str(value)
    style_kwargs = {"stylename": _ODS_PERCENT_CELL_STYLE_NAME} if percent else {}

    if formula:
        # Formula cells carry a cached float value of 0.0 and the formula
        # attribute; the displayed text is left empty.
        cell = TableCell(valuetype="float", value=0.0, **style_kwargs)
        cell.setAttribute("formula", formula)
        cell.addElement(P(text=text))
    elif isinstance(value, bool):
        cell = TableCell(valuetype="boolean", value="true" if value else "false", **style_kwargs)
        cell.addElement(P(text=text))
    elif isinstance(value, (int, float)):
        cell = TableCell(valuetype="float", value=float(value), **style_kwargs)
        cell.addElement(P(text=text))
    else:
        cell = TableCell(valuetype="string", **style_kwargs)
        cell.addElement(P(text=text))
    return cell


def _ods_add_row(table, values, start_col=0, formulas=None, percent_cols=None):
    """Append a row to an ODS ``Table`` with leading empty cells and formulas.

    Args:
        table: odfpy ``Table`` instance.
        values: iterable of cell values placed starting at ``start_col``.
        start_col: 0-based column index where ``values`` begin.
        formulas: dict mapping absolute 0-based column index to an ODS formula
            string (``of:=...``).  The matching cell is created as a formula
            cell with the corresponding ``values`` entry as display text.
        percent_cols: set of absolute 0-based column indices to render with
            the percent cell style.
    """
    from odf.table import TableRow

    formulas = formulas or {}
    percent_cols = percent_cols or set()
    row = TableRow()
    end_col = max(len(values) + start_col, max(formulas, default=-1) + 1)
    for col_idx in range(end_col):
        if start_col <= col_idx < start_col + len(values):
            value = values[col_idx - start_col]
        else:
            value = ""
        cell = _ods_create_cell(
            value, formula=formulas.get(col_idx), percent=col_idx in percent_cols,
        )
        row.addElement(cell)
    table.addElement(row)
    return row


def write_ods_sheet(document, result: AllocationResult, config: AllocationConfig,
                    report_date, generator_version: str):
    """Write the ``Allocation`` sheet into an existing ``odfpy`` document.

    Layout mirrors ``write_xlsx_sheet`` exactly (RFC allocation_sheet.md §5 and
    ticket E11-S10):
      - Row 1: title, report date, generator version.
      - Row 2: visible note that the sheet is unprotected.
      - Rows 3-13: config block (parameter name, editable value, shipped default).
      - Rows 14-17: summary formulas (selected count, gross exposure %, enabled
        count, EV total, and the gross scale factor in $E$14).
      - Row 18: blank spacer row.
      - Row 19: instruction line.
      - Row 20: header row A-AO (Model in AO is a trailing plain value).
      - Rows 21..N: one row per candidate in ``result.rows`` in the order given.
      - Below the data: Section B cluster-exposure table and Section C rejected
        signals table.

    No sheet protection is applied because odfpy protection support is
    unreliable (RFC §4.6); the unprotected state is surfaced to the user via
    the row-2 note.

    The function performs no disk I/O; it mutates the provided in-memory
    ``OpenDocumentSpreadsheet``.
    """
    from odf import opendocument
    from odf.table import Table

    if not isinstance(document, opendocument.OpenDocument):
        raise TypeError("document must be an odfpy OpenDocumentSpreadsheet")
    if document.mimetype != "application/vnd.oasis.opendocument.spreadsheet":
        raise TypeError("document must be an odfpy OpenDocumentSpreadsheet")

    _ods_ensure_percent_style(document)

    table = Table(name="Allocation")
    document.spreadsheet.addElement(table)

    # -------------------------------------------------------------------------
    # Row 1: title line
    # -------------------------------------------------------------------------
    _ods_add_row(
        table,
        ["Portfolio Allocation", _empty_if_none(report_date), f"generator {generator_version}"],
    )

    # -------------------------------------------------------------------------
    # Row 2: unprotected-sheet note (ODS protection is unreliable)
    # -------------------------------------------------------------------------
    _ods_add_row(
        table,
        ["Note: this sheet is intentionally unprotected (ODS protection support is unreliable)."],
    )

    # -------------------------------------------------------------------------
    # Rows 3-13: config block
    # -------------------------------------------------------------------------
    # Parameter names in column D, editable values in column E, shipped defaults
    # in column F.
    default_config = AllocationConfig()
    for label, attr in _CONFIG_BLOCK:
        is_pct = "pct" in label
        value = getattr(config, attr)
        if attr == "cluster_map":
            value = str(value) if value else ""
        elif is_pct:
            value = _pct_to_frac(value)
        default_value = getattr(default_config, attr)
        if attr == "cluster_map":
            default_value = str(default_value) if default_value else ""
        elif is_pct:
            default_value = _pct_to_frac(default_value)
        _ods_add_row(
            table,
            [label, _empty_if_none(value), _empty_if_none(default_value)],
            start_col=3,
            percent_cols={4, 5} if is_pct else None,
        )

    # -------------------------------------------------------------------------
    # Rows 14-17: summary block
    # -------------------------------------------------------------------------
    _ods_add_row(
        table,
        ["Selected count", ""],
        formulas={
            3: _xlsx_formula_to_ods(_SUMMARY_FORMULAS["selected_count"]),
            4: render_formula("gross_scale", 14, "ods"),
        },
    )
    _ods_add_row(
        table,
        ["Gross exposure %", ""],
        formulas={3: _xlsx_formula_to_ods(_SUMMARY_FORMULAS["gross_exposure"])},
        percent_cols={3},
    )
    _ods_add_row(
        table,
        ["Enabled count", ""],
        formulas={3: _xlsx_formula_to_ods(_SUMMARY_FORMULAS["enabled_count"])},
    )

    # -------------------------------------------------------------------------
    # Row 17: EV total summary row (mirrors XLSX layout)
    # -------------------------------------------------------------------------
    _ods_add_row(
        table,
        ["EV total", ""],
        formulas={3: _xlsx_formula_to_ods(_SUMMARY_FORMULAS["ev_total_sum"])},
        percent_cols={3},
    )

    # -------------------------------------------------------------------------
    # Row 18: blank spacer row (intentionally left empty)
    # -------------------------------------------------------------------------
    _ods_add_row(table, [""])

    # -------------------------------------------------------------------------
    # Row 19: instruction line
    # -------------------------------------------------------------------------
    _ods_add_row(
        table,
        [
            "Edit only the config values (column E), the per-row input column "
            "(column O), and the Enabled column (column A). All other cells are computed."
        ],
    )

    # -------------------------------------------------------------------------
    # Row 20: header row
    # -------------------------------------------------------------------------
    headers = (
        _STATIC_HEADERS
        + [_FORMULA_HEADERS[col] for col in _FORMULA_COLS[:8]]
        + _HELPER_HEADERS
        + ["Model"]
    )
    _ods_add_row(table, headers)

    # -------------------------------------------------------------------------
    # Rows 21..N: candidate rows, sorted by EV total descending
    # -------------------------------------------------------------------------
    def _cluster_for_ticker(ticker: str) -> str:
        return config.cluster_map.get(ticker, ticker)

    # 0-based column indices of the static percent-labeled columns (I, J, O).
    static_percent_cols = {
        _xlsx_column_letter_to_index(c) - 1 for c in ("I", "J", "O")
    }
    formula_percent_cols = {
        _xlsx_column_letter_to_index(c) - 1 for c in _PERCENT_COLS
        if c not in ("I", "J", "O")
    }

    sorted_rows = _sorted_for_sheet(result.rows)

    for row_offset, row_data in enumerate(sorted_rows):
        excel_row = _DATA_START_ROW + 1 + row_offset
        derived = row_data.get("derived", {}) or {}
        ticker = row_data.get("ticker", "")
        direction = row_data.get("direction", "")
        enabled_value = "true" if row_data.get("status") == "SELECTED" else "false"

        static_values = [
            enabled_value,
            ticker,
            _cluster_for_ticker(ticker),
            row_data.get("strategy", ""),
            direction.capitalize() if isinstance(direction, str) else "",
            row_data.get("entry"),
            row_data.get("stop"),
            row_data.get("target"),
            _pct_to_frac(derived.get("risk_pct")),
            _pct_to_frac(derived.get("reward_pct")),
            derived.get("b"),
            row_data.get("n"),
            row_data.get("base_win_rate"),
            derived.get("p_shrunk"),
            _pct_to_frac(row_data.get("ev_pct")),
        ]

        # Visible + helper formula columns P..AN are formula cells.
        formulas = {}
        for col_letter in _FORMULA_COLS:
            col_idx = _xlsx_column_letter_to_index(col_letter) - 1
            formulas[col_idx] = render_formula(col_letter, excel_row, "ods")

        # Trailing plain-value Model column, appended after all A-AN columns.
        values = static_values + [""] * len(_FORMULA_COLS) + [row_data.get("model") or ""]
        percent_cols = static_percent_cols | formula_percent_cols
        _ods_add_row(table, values, formulas=formulas, percent_cols=percent_cols)

    # -------------------------------------------------------------------------
    # Autofilter: table:database-range over header + data rows (A20:AN<end>)
    # -------------------------------------------------------------------------
    from odf.table import DatabaseRange, DatabaseRanges

    data_end_row = _DATA_START_ROW + len(sorted_rows)
    db_ranges = document.spreadsheet.getElementsByType(DatabaseRanges)
    if db_ranges:
        db_ranges_elem = db_ranges[0]
    else:
        db_ranges_elem = DatabaseRanges()
        document.spreadsheet.addElement(db_ranges_elem)
    db_ranges_elem.addElement(
        DatabaseRange(
            name="AllocationFilter",
            targetrangeaddress=(
                f"'Allocation'.A{_HEADER_ROW}:'Allocation'.{_MODEL_COL}{data_end_row}"
            ),
            displayfilterbuttons="true",
        )
    )

    # -------------------------------------------------------------------------
    # Section B: cluster exposure table
    # -------------------------------------------------------------------------
    _ods_add_row(table, [""])  # blank separator row
    _ods_add_row(
        table,
        ["Cluster", "Positions", "Gross %", "Cap %", "Capped?"],
    )

    selected_rows = [r for r in result.rows if r.get("status") == "SELECTED"]
    clusters = sorted(set(config.cluster_map.values())) if config.cluster_map else []
    for cluster in clusters:
        cluster_rows = [
            r for r in selected_rows
            if _cluster_for_ticker(r.get("ticker", "")) == cluster
        ]
        positions = len(cluster_rows)
        gross = sum(r.get("alloc", 0.0) for r in cluster_rows)
        capped = any("CLUSTER_CAPPED" in r.get("flags", []) for r in cluster_rows)
        _ods_add_row(
            table,
            [
                cluster, positions,
                _pct_to_frac(gross), _pct_to_frac(config.max_cluster_pct),
                "yes" if capped else "no",
            ],
            percent_cols={2, 3},
        )

    # -------------------------------------------------------------------------
    # Section C: rejected signals (compact audit trail)
    # -------------------------------------------------------------------------
    _ods_add_row(table, ["Ticker", "Strategy", "Dir", "Score", "Reason"])

    rejected = [r for r in result.rows if r.get("status") != "SELECTED"]
    rejected.sort(
        key=lambda r: (
            r.get("status", ""),
            -(r.get("derived", {}) or {}).get("score", float("-inf")),
        )
    )
    for r in rejected:
        direction = r.get("direction", "")
        score = (r.get("derived", {}) or {}).get("score")
        _ods_add_row(
            table,
            [
                r.get("ticker", ""),
                r.get("strategy", ""),
                direction.capitalize() if isinstance(direction, str) else "",
                _empty_if_none(score),
                r.get("status", ""),
            ],
        )


# =============================================================================
# E11-S11: Markdown Section Writer
# =============================================================================

def write_md_section(result: AllocationResult, config: AllocationConfig) -> str:
    """Render the RFC §6 Markdown "Portfolio Allocation" section.

    Returns a static snapshot of ``result`` as a markdown string, including the
    config summary, selection summary, selected-position table, cluster
    exposure line, and compact rejection-count summary.

    Args:
        result: AllocationResult from ``allocate()``.
        config: AllocationConfig used to produce ``result``.

    Returns:
        Markdown string ready to append to a signals report.
    """
    # Heading
    lines = ["## Portfolio Allocation", ""]

    # Config summary line (RFC §6 example format)
    lines.append(
        f"Config: n0={config.n0:g} min_n={config.min_n:g} "
        f"cost={config.round_trip_cost_pct:g}% kelly_mult={config.kelly_mult:g} "
        f"top_k={config.top_k:g} max_pos={config.max_pos_pct:g}% "
        f"max_cluster={config.max_cluster_pct:g}% gross_cap={config.gross_cap_pct:g}%"
    )
    lines.append("")

    # Selection summary
    total_signals = len(result.rows)
    ev_total_pct = sum(
        (row.get("ev_pct") or 0) / 100 * (row.get("alloc") or 0) / 100
        for row in result.rows
    ) * 100
    lines.append(
        f"Selected {result.selected_count} of {total_signals} signals. "
        f"Gross exposure: {result.gross_exposure_pct:.1f}%. "
        f"EV total: {ev_total_pct:.2f}%."
    )
    lines.append("")

    # Selected-position table via kairos_signals.format_table
    headers = ["Ticker", "Dir", "Strategy", "Entry", "Stop", "Target", "EV net", "Score", "Alloc", "Model"]
    align = ["l", "l", "l", "r", "r", "r", "r", "r", "r", "l"]

    selected_rows = [row for row in result.rows if row.get("status") == "SELECTED"]
    table_rows = []
    for row in selected_rows:
        direction = row.get("direction", "")
        derived = row.get("derived", {}) or {}
        ev_net = derived.get("ev_net")
        score = derived.get("score")
        alloc = row.get("alloc", 0.0)

        entry = row.get("entry")
        stop = row.get("stop")
        target = row.get("target")

        table_rows.append({
            "Ticker": row.get("ticker", ""),
            "Dir": direction.capitalize() if isinstance(direction, str) else "",
            "Strategy": row.get("strategy", ""),
            "Entry": f"{entry:.2f}" if entry is not None else "",
            "Stop": f"{stop:.2f}" if stop is not None else "",
            "Target": f"{target:.2f}" if target is not None else "",
            "EV net": f"{ev_net:.2f}%" if ev_net is not None else "",
            "Score": f"{score:.2f}" if score is not None else "",
            "Alloc": f"{alloc:.1f}%",
            "Model": row.get("model") or "",
        })

    table_lines = format_table(headers, table_rows, align)
    lines.extend(table_lines)
    lines.append("")

    # Cluster exposure line: sum alloc per cluster among selected rows
    cluster_sums: dict[str, float] = {}
    for row in selected_rows:
        ticker = row.get("ticker", "")
        cluster = config.cluster_map.get(ticker, ticker)
        cluster_sums[cluster] = cluster_sums.get(cluster, 0.0) + row.get("alloc", 0.0)

    if cluster_sums:
        sorted_clusters = sorted(cluster_sums.items(), key=lambda kv: (-kv[1], kv[0]))
        cluster_parts = [f"{cluster} {total:.1f}%" for cluster, total in sorted_clusters]
        lines.append(f"Cluster exposure: {', '.join(cluster_parts)}")
    else:
        lines.append("Cluster exposure: none")
    lines.append("")

    # Rejection-count summary line
    non_selected_count = total_signals - result.selected_count
    if result.rejection_counts:
        sorted_reasons = sorted(
            result.rejection_counts.items(),
            key=lambda kv: (-kv[1], kv[0]),
        )
        reason_parts = [f"{reason} {count}" for reason, count in sorted_reasons]
        lines.append(f"Rejected: {non_selected_count} total -- {', '.join(reason_parts)}")
    else:
        lines.append(f"Rejected: {non_selected_count} total --")

    return "\n".join(lines)
